"""
Part 4: Monthly Risk Report Template
======================================
A professional monthly risk report for Fund Alpha, modelled on what
institutional risk teams actually produce.

Picks up returns_df from the notebook if available.

SECTIONS:
  1. Executive Summary — NAV, month/YTD/inception performance
  2. Risk Metrics Snapshot — current VaR, vol, beta, Sharpe + change vs prior month
  3. Exposure Breakdown — asset class allocations
  4. Drawdown Monitor — current drawdown, time in drawdown, historical max
  5. Stress Testing — scenario analysis (2008, 2020, rate shock, etc.)
  6. Limit Monitoring — traffic light system for risk limits
  7. Risk Commentary — template section for analyst notes
"""

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.utils import get_column_letter
from datetime import datetime

# ── Configuration ─────────────────────────────────────────────────────────
RISK_FREE_RATE = 0.045
BENCHMARK_NAME = "Global Equity Index"
FUND_NAME = "Fund Alpha"

# Risk limits (these would be set by the fund's risk committee)
LIMITS = {
    "Max Drawdown Limit": -0.20,          # -20% — fund gets reviewed if breached
    "Monthly VaR Limit (95%)": -0.08,     # -8% — worst expected monthly loss
    "Annualised Volatility Limit": 0.25,  # 25% — max acceptable vol
    "Beta Limit (vs Equity)": 0.50,       # 0.5 — should stay market-neutral-ish
    "Concentration Limit (single asset)": 0.25,  # 25% max in any one asset class
}

# Simulated exposure data (not available from Yahoo Finance)
EXPOSURES = {
    "Equities": {"long": 0.35, "short": 0.20, "weight": 0.15},
    "Fixed Income": {"long": 0.40, "short": 0.15, "weight": 0.25},
    "Commodities": {"long": 0.25, "short": 0.10, "weight": 0.15},
    "FX": {"long": 0.30, "short": 0.25, "weight": 0.05},
    "Rates": {"long": 0.20, "short": 0.10, "weight": 0.10},
}

# Stress scenarios (historical and hypothetical)
SCENARIOS = {
    "2008 GFC (Sep-Nov)": -0.08,
    "2020 COVID Crash (Mar)": -0.05,
    "2022 Rate Shock (H1)": 0.12,
    "Equity -20% Shock": -0.04,
    "Bond -10% Shock": -0.03,
    "Commodity +30% Spike": 0.06,
    "USD +10% Rally": 0.02,
    "Vol Spike (VIX to 40)": 0.03,
}

# ── Pick up data from notebook ────────────────────────────────────────────
try:
    _ = returns_df  # noqa: F821
    if FUND_NAME in returns_df.columns:
        fund_returns = returns_df[FUND_NAME].copy()
        bench_returns = returns_df[BENCHMARK_NAME].copy() if BENCHMARK_NAME in returns_df.columns else None
        print(f"✅ Using REAL data from notebook ({len(fund_returns)} months)")
    else:
        raise NameError("Fund not found")
except NameError:
    print("⚠️  No data loaded — using SIMULATED data")
    np.random.seed(42)
    dates = pd.date_range("2019-01-01", "2025-12-31", freq="ME")
    n = len(dates)
    market = np.random.normal(0, 0.04, n)
    fund_returns = pd.Series(
        0.3 * market + np.random.normal(0.005, 0.035, n), index=dates, name=FUND_NAME)
    bench_returns = pd.Series(
        market + np.random.normal(0.005, 0.01, n), index=dates, name=BENCHMARK_NAME)

N = len(fund_returns)
dates = fund_returns.index

# ── Calculate all metrics ─────────────────────────────────────────────────
# Current period metrics (full history)
ann_ret = (1 + fund_returns).prod() ** (12 / N) - 1
ann_vol = fund_returns.std() * np.sqrt(12)
sharpe = (ann_ret - RISK_FREE_RATE) / ann_vol if ann_vol != 0 else 0
sortino_dv = fund_returns[fund_returns < 0].std() * np.sqrt(12)
sortino = (ann_ret - RISK_FREE_RATE) / sortino_dv if sortino_dv != 0 else 0

cum = (1 + fund_returns).cumprod()
peak = cum.cummax()
dd = (cum - peak) / peak
max_dd = dd.min()
current_dd = dd.iloc[-1]

var_95 = np.percentile(fund_returns, 5)
cvar_95 = fund_returns[fund_returns <= var_95].mean()

if bench_returns is not None:
    cov = np.cov(fund_returns, bench_returns)[0, 1]
    var_b = np.var(bench_returns, ddof=1)
    beta = cov / var_b if var_b != 0 else 0
else:
    beta = 0

# Prior month metrics (excluding last month) for comparison
if N > 12:
    prev_returns = fund_returns.iloc[:-1]
    prev_ann_ret = (1 + prev_returns).prod() ** (12 / len(prev_returns)) - 1
    prev_ann_vol = prev_returns.std() * np.sqrt(12)
    prev_sharpe = (prev_ann_ret - RISK_FREE_RATE) / prev_ann_vol if prev_ann_vol != 0 else 0
    prev_var = np.percentile(prev_returns, 5)
    prev_cum = (1 + prev_returns).cumprod()
    prev_peak = prev_cum.cummax()
    prev_dd = ((prev_cum - prev_peak) / prev_peak).min()
else:
    prev_ann_ret = prev_ann_vol = prev_sharpe = prev_var = prev_dd = 0

# Monthly returns
latest_month_ret = fund_returns.iloc[-1]
ytd_months = fund_returns[fund_returns.index.year == dates[-1].year]
ytd_ret = (1 + ytd_months).prod() - 1

# Rolling metrics
rolling_vol_12m = fund_returns.rolling(12).std() * np.sqrt(12)
rolling_sharpe_12m = fund_returns.rolling(12).apply(
    lambda x: ((1 + x).prod() ** (12/len(x)) - 1 - RISK_FREE_RATE) / (x.std() * np.sqrt(12))
    if x.std() != 0 else 0, raw=False)

# Time in drawdown
dd_periods = (dd < 0).astype(int)
current_dd_length = 0
for v in reversed(dd_periods.values):
    if v == 1:
        current_dd_length += 1
    else:
        break

# NAV simulation (growth of 1000)
nav = cum * 1000
current_nav = nav.iloc[-1]
peak_nav = nav.max()


# ── Styles ────────────────────────────────────────────────────────────────
DARK_BLUE = "1B2A4A"
MED_BLUE = "2E5090"
LIGHT_BLUE = "D6E4F0"
WHITE = "FFFFFF"
LIGHT_GREY = "F2F2F2"
GREEN = "27AE60"
RED = "E74C3C"
AMBER = "F39C12"
DARK_GREEN = "1E8449"

hdr_font = Font(name="Arial", bold=True, color=WHITE, size=11)
hdr_fill = PatternFill("solid", fgColor=DARK_BLUE)
hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
data_font = Font(name="Arial", size=10)
label_font = Font(name="Arial", bold=True, size=10, color=DARK_BLUE)
thin_border = Border(bottom=Side(style="thin", color="D0D0D0"))
pct_fmt = '0.0%'
pct_fmt2 = '0.00%'
ratio_fmt = '0.00'
nav_fmt = '#,##0'

def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cl = ws.cell(row=row, column=c)
        cl.font, cl.fill, cl.alignment = hdr_font, hdr_fill, hdr_align

def style_row(ws, row, cols, alt=False):
    fill = PatternFill("solid", fgColor=LIGHT_GREY) if alt else PatternFill("solid", fgColor=WHITE)
    for c in range(1, cols + 1):
        cl = ws.cell(row=row, column=c)
        cl.font = data_font
        cl.fill = fill
        cl.border = thin_border

def title_banner(ws, text, end_col="H", tab_color=DARK_BLUE):
    ws.merge_cells(f"A1:{end_col}1")
    ws["A1"] = text
    ws["A1"].font = Font(name="Arial", bold=True, color=WHITE, size=14)
    ws["A1"].fill = PatternFill("solid", fgColor=DARK_BLUE)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 38
    ws.sheet_properties.tabColor = tab_color

def traffic_light(ws, row, col, status):
    """Apply green/amber/red fill based on status."""
    colours = {
        "GREEN": (DARK_GREEN, WHITE), "AMBER": (AMBER, WHITE), "RED": (RED, WHITE)
    }
    fill_col, font_col = colours.get(status, (WHITE, "000000"))
    cell = ws.cell(row=row, column=col)
    cell.value = status
    cell.fill = PatternFill("solid", fgColor=fill_col)
    cell.font = Font(name="Arial", bold=True, size=10, color=font_col)
    cell.alignment = Alignment(horizontal="center")

def set_widths(ws, cols, default=15):
    for c in range(1, cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = default
    ws.column_dimensions["A"].width = 30


wb = Workbook()
report_date = dates[-1].strftime("%B %Y")


# ══════════════════════════════════════════════════════════════════════════
# SHEET 1: Executive Summary
# ══════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Executive Summary"
title_banner(ws1, f"MONTHLY RISK REPORT — {FUND_NAME.upper()}", end_col="F")

# Report info bar
ws1.merge_cells("A2:F2")
ws1["A2"] = f"Report Date: {report_date}  |  Benchmark: {BENCHMARK_NAME}  |  Risk-Free Rate: {RISK_FREE_RATE:.1%}"
ws1["A2"].font = Font(name="Arial", italic=True, color=DARK_BLUE, size=9)
ws1["A2"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)
ws1["A2"].alignment = Alignment(horizontal="center")

# NAV & Performance
ws1["A4"] = "FUND OVERVIEW"
ws1["A4"].font = Font(name="Arial", bold=True, size=13, color=DARK_BLUE)

overview_data = [
    ("Fund Name", FUND_NAME, None),
    ("Strategy", "Systematic Trend-Following / Managed Futures", None),
    ("Benchmark", BENCHMARK_NAME, None),
    ("Inception Date", dates[0].strftime("%B %Y"), None),
    ("Report Period End", dates[-1].strftime("%d %B %Y"), None),
    ("", "", None),
    ("PERFORMANCE", "", None),
    ("Current NAV (per $1,000 invested)", round(current_nav, 2), nav_fmt),
    ("Peak NAV", round(float(peak_nav), 2), nav_fmt),
    ("Latest Month Return", round(latest_month_ret, 4), pct_fmt),
    ("Year-to-Date Return", round(ytd_ret, 4), pct_fmt),
    ("Annualised Return (inception)", round(ann_ret, 4), pct_fmt),
    ("", "", None),
    ("RISK SNAPSHOT", "", None),
    ("Annualised Volatility", round(ann_vol, 4), pct_fmt),
    ("Sharpe Ratio", round(sharpe, 2), ratio_fmt),
    ("Maximum Drawdown", round(max_dd, 4), pct_fmt),
    ("Current Drawdown", round(current_dd, 4), pct_fmt),
    ("VaR (95%, 1-Month)", round(var_95, 4), pct_fmt),
    ("Beta vs Benchmark", round(beta, 2), ratio_fmt),
]

for i, (label, value, fmt) in enumerate(overview_data):
    r = 6 + i
    ws1.cell(row=r, column=1, value=label)
    cell_v = ws1.cell(row=r, column=2, value=value)
    if fmt:
        cell_v.number_format = fmt

    if label in ("PERFORMANCE", "RISK SNAPSHOT", "FUND OVERVIEW"):
        ws1.cell(row=r, column=1).font = Font(name="Arial", bold=True, size=11, color=DARK_BLUE)
    elif label == "":
        pass
    else:
        ws1.cell(row=r, column=1).font = label_font
        cell_v.font = Font(name="Arial", size=10, bold=True)

set_widths(ws1, 6, default=18)
ws1.column_dimensions["A"].width = 36
ws1.column_dimensions["B"].width = 28


# ══════════════════════════════════════════════════════════════════════════
# SHEET 2: Risk Metrics Detail (with month-over-month changes)
# ══════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Risk Metrics")
title_banner(ws2, "RISK METRICS — DETAILED VIEW", end_col="E", tab_color=MED_BLUE)

ws2.merge_cells("A2:E2")
ws2["A2"] = "Shows current values vs prior month. Green arrow = improving, Red = deteriorating."
ws2["A2"].font = Font(name="Arial", italic=True, size=9, color=DARK_BLUE)
ws2["A2"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)

row_h = 4
headers = ["Metric", "Current", "Prior Month", "Change", "Direction"]
for c, h in enumerate(headers, 1):
    ws2.cell(row=row_h, column=c, value=h)
style_header(ws2, row_h, len(headers))

# For each metric: (name, current, prior, higher_is_better)
risk_detail = [
    ("Annualised Return", ann_ret, prev_ann_ret, True, pct_fmt),
    ("Annualised Volatility", ann_vol, prev_ann_vol, False, pct_fmt),
    ("Sharpe Ratio", sharpe, prev_sharpe, True, ratio_fmt),
    ("Sortino Ratio", sortino, 0, True, ratio_fmt),
    ("Maximum Drawdown", max_dd, prev_dd, False, pct_fmt),
    ("Current Drawdown", current_dd, 0, False, pct_fmt),
    ("VaR (95%, 1M)", var_95, prev_var, False, pct_fmt),
    ("CVaR (95%, 1M)", cvar_95, 0, False, pct_fmt),
    ("Beta vs Benchmark", beta, 0, False, ratio_fmt),
    ("Win Rate", (fund_returns > 0).mean(), 0, True, pct_fmt),
]

for i, (name, current, prior, higher_better, fmt) in enumerate(risk_detail):
    r = row_h + 1 + i
    ws2.cell(row=r, column=1, value=name).font = label_font
    ws2.cell(row=r, column=2, value=round(current, 4)).number_format = fmt
    ws2.cell(row=r, column=3, value=round(prior, 4)).number_format = fmt

    change = current - prior
    ws2.cell(row=r, column=4, value=round(change, 4)).number_format = fmt

    # Direction arrow
    if prior != 0 and change != 0:
        if (higher_better and change > 0) or (not higher_better and change < 0):
            ws2.cell(row=r, column=5, value="▲ Improving")
            ws2.cell(row=r, column=5).font = Font(name="Arial", bold=True, color=GREEN, size=10)
        elif (higher_better and change < 0) or (not higher_better and change > 0):
            ws2.cell(row=r, column=5, value="▼ Deteriorating")
            ws2.cell(row=r, column=5).font = Font(name="Arial", bold=True, color=RED, size=10)
        else:
            ws2.cell(row=r, column=5, value="— Unchanged")
    else:
        ws2.cell(row=r, column=5, value="—")

    style_row(ws2, r, len(headers), alt=(i % 2 == 0))

# Rolling volatility chart
vol_chart_row = row_h + len(risk_detail) + 3
ws2.cell(row=vol_chart_row, column=1, value="ROLLING 12-MONTH VOLATILITY").font = Font(
    name="Arial", bold=True, size=12, color=DARK_BLUE)

vr = vol_chart_row + 1
ws2.cell(row=vr, column=1, value="Date")
ws2.cell(row=vr, column=2, value="Rolling 12M Vol")
style_header(ws2, vr, 2)

rv = rolling_vol_12m.dropna()
for i, (dt, val) in enumerate(rv.items()):
    r = vr + 1 + i
    ws2.cell(row=r, column=1, value=dt).number_format = "MMM-YY"
    ws2.cell(row=r, column=2, value=round(val, 4)).number_format = pct_fmt

vol_chart = LineChart()
vol_chart.title = "Rolling 12-Month Volatility"
vol_chart.y_axis.numFmt = '0%'
vol_chart.y_axis.title = "Annualised Volatility"
vol_chart.style = 10
vol_chart.width = 28
vol_chart.height = 14
vol_ref = Reference(ws2, min_col=2, min_row=vr, max_row=vr + len(rv))
vol_chart.add_data(vol_ref, titles_from_data=True)
vol_chart.series[0].graphicalProperties.line.solidFill = MED_BLUE
vol_chart.series[0].graphicalProperties.line.width = 22000
cats = Reference(ws2, min_col=1, min_row=vr + 1, max_row=vr + len(rv))
vol_chart.set_categories(cats)
vol_chart.legend = None
ws2.add_chart(vol_chart, f"D{vol_chart_row}")

set_widths(ws2, len(headers), default=18)
ws2.column_dimensions["A"].width = 28
ws2.column_dimensions["E"].width = 20


# ══════════════════════════════════════════════════════════════════════════
# SHEET 3: Exposure Breakdown
# ══════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Exposure Breakdown")
title_banner(ws3, "PORTFOLIO EXPOSURE BREAKDOWN", end_col="F", tab_color="27AE60")

ws3.merge_cells("A2:F2")
ws3["A2"] = "Estimated asset class exposures. In practice, these come from the portfolio management system daily."
ws3["A2"].font = Font(name="Arial", italic=True, size=9, color=DARK_BLUE)
ws3["A2"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)

row_h = 4
exp_headers = ["Asset Class", "Long Exposure", "Short Exposure", "Net Exposure", "Gross Exposure", "Portfolio Weight"]
for c, h in enumerate(exp_headers, 1):
    ws3.cell(row=row_h, column=c, value=h)
style_header(ws3, row_h, len(exp_headers))

total_long = total_short = total_net = total_gross = total_weight = 0
for i, (asset, exp) in enumerate(EXPOSURES.items()):
    r = row_h + 1 + i
    ws3.cell(row=r, column=1, value=asset).font = label_font
    ws3.cell(row=r, column=2, value=exp["long"]).number_format = pct_fmt
    ws3.cell(row=r, column=3, value=exp["short"]).number_format = pct_fmt
    net = exp["long"] - exp["short"]
    gross = exp["long"] + exp["short"]
    ws3.cell(row=r, column=4, value=round(net, 4)).number_format = pct_fmt
    ws3.cell(row=r, column=5, value=round(gross, 4)).number_format = pct_fmt
    ws3.cell(row=r, column=6, value=exp["weight"]).number_format = pct_fmt

    # Colour net exposure
    if net > 0:
        ws3.cell(row=r, column=4).font = Font(name="Arial", size=10, color=GREEN)
    elif net < 0:
        ws3.cell(row=r, column=4).font = Font(name="Arial", size=10, color=RED)

    total_long += exp["long"]
    total_short += exp["short"]
    total_net += net
    total_gross += gross
    total_weight += exp["weight"]

    for c in [2, 3, 4, 5, 6]:
        ws3.cell(row=r, column=c).font = Font(name="Arial", size=10, color="0000FF")
    style_row(ws3, r, len(exp_headers), alt=(i % 2 == 0))

# Totals
tr = row_h + 1 + len(EXPOSURES)
ws3.cell(row=tr, column=1, value="TOTAL").font = Font(name="Arial", bold=True, size=10)
for c, val in [(2, total_long), (3, total_short), (4, total_net), (5, total_gross), (6, total_weight)]:
    ws3.cell(row=tr, column=c, value=round(val, 4)).number_format = pct_fmt
    ws3.cell(row=tr, column=c).font = Font(name="Arial", bold=True, size=10)
    ws3.cell(row=tr, column=c).border = Border(
        top=Side(style="double"), bottom=Side(style="double"))

# Key definitions
def_row = tr + 3
defs = [
    ("KEY DEFINITIONS:", ""),
    ("Long Exposure:", "Notional value of long positions as % of NAV"),
    ("Short Exposure:", "Notional value of short positions as % of NAV"),
    ("Net Exposure:", "Long minus Short — shows directional bias"),
    ("Gross Exposure:", "Long plus Short — shows total market engagement (leverage indicator)"),
    ("Portfolio Weight:", "Risk-budgeted allocation to each asset class"),
]
for i, (label, desc) in enumerate(defs):
    r = def_row + i
    ws3.cell(row=r, column=1, value=label).font = Font(
        name="Arial", bold=True, size=10, color=DARK_BLUE if i == 0 else "333333")
    ws3.cell(row=r, column=2, value=desc).font = Font(name="Arial", size=9, italic=True)

# Exposure bar chart
exp_chart_row = def_row + len(defs) + 2
bar = BarChart()
bar.type = "col"
bar.title = "Net Exposure by Asset Class"
bar.y_axis.numFmt = '0%'
bar.style = 10
bar.width = 22
bar.height = 14

data_ref = Reference(ws3, min_col=4, min_row=row_h, max_row=row_h + len(EXPOSURES))
cats_ref = Reference(ws3, min_col=1, min_row=row_h + 1, max_row=row_h + len(EXPOSURES))
bar.add_data(data_ref, titles_from_data=True)
bar.set_categories(cats_ref)
bar.series[0].graphicalProperties.solidFill = MED_BLUE
bar.legend = None
ws3.add_chart(bar, f"A{exp_chart_row}")

set_widths(ws3, len(exp_headers), default=18)
ws3.column_dimensions["A"].width = 22


# ══════════════════════════════════════════════════════════════════════════
# SHEET 4: Drawdown Monitor
# ══════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Drawdown Monitor")
title_banner(ws4, "DRAWDOWN MONITORING", end_col="D", tab_color=RED)

ws4["A3"] = "CURRENT STATUS"
ws4["A3"].font = Font(name="Arial", bold=True, size=13, color=DARK_BLUE)

dd_summary = [
    ("Current Drawdown", current_dd, pct_fmt),
    ("Maximum Drawdown (all-time)", max_dd, pct_fmt),
    ("Current NAV", current_nav, '$#,##0.00'),
    ("Peak NAV", float(peak_nav), '$#,##0.00'),
    ("Months in Current Drawdown", current_dd_length, '0'),
    ("Drawdown Limit", LIMITS["Max Drawdown Limit"], pct_fmt),
    ("Limit Headroom", current_dd - LIMITS["Max Drawdown Limit"], pct_fmt),
]

for i, (label, value, fmt) in enumerate(dd_summary):
    r = 5 + i
    ws4.cell(row=r, column=1, value=label).font = label_font
    cell = ws4.cell(row=r, column=2, value=round(value, 4) if isinstance(value, float) else value)
    cell.number_format = fmt
    cell.font = Font(name="Arial", bold=True, size=11)
    if label == "Limit Headroom":
        cell.font = Font(name="Arial", bold=True, size=11,
                        color=GREEN if value > 0 else RED)

# Drawdown status
status_row = 5 + len(dd_summary) + 1
ws4.cell(row=status_row, column=1, value="LIMIT STATUS:").font = Font(
    name="Arial", bold=True, size=11, color=DARK_BLUE)
if current_dd > LIMITS["Max Drawdown Limit"]:
    traffic_light(ws4, status_row, 2, "GREEN")
elif current_dd > LIMITS["Max Drawdown Limit"] * 0.8:
    traffic_light(ws4, status_row, 2, "AMBER")
else:
    traffic_light(ws4, status_row, 2, "RED")

# Drawdown time series
dd_chart_row = status_row + 3
ws4.cell(row=dd_chart_row, column=1, value="DRAWDOWN HISTORY").font = Font(
    name="Arial", bold=True, size=12, color=DARK_BLUE)

dr = dd_chart_row + 1
ws4.cell(row=dr, column=1, value="Date")
ws4.cell(row=dr, column=2, value="Drawdown")
style_header(ws4, dr, 2)

for i, (dt, val) in enumerate(dd.items()):
    r = dr + 1 + i
    ws4.cell(row=r, column=1, value=dt).number_format = "MMM-YY"
    ws4.cell(row=r, column=2, value=round(val, 4)).number_format = pct_fmt

dd_chart = LineChart()
dd_chart.title = "Underwater Chart"
dd_chart.y_axis.numFmt = '0%'
dd_chart.y_axis.title = "Drawdown"
dd_chart.style = 10
dd_chart.width = 28
dd_chart.height = 14
dd_ref = Reference(ws4, min_col=2, min_row=dr, max_row=dr + N)
dd_chart.add_data(dd_ref, titles_from_data=True)
dd_chart.series[0].graphicalProperties.line.solidFill = RED
dd_chart.series[0].graphicalProperties.line.width = 22000
cats = Reference(ws4, min_col=1, min_row=dr + 1, max_row=dr + N)
dd_chart.set_categories(cats)
dd_chart.legend = None
ws4.add_chart(dd_chart, f"D{dd_chart_row}")

set_widths(ws4, 4, default=20)
ws4.column_dimensions["A"].width = 36


# ══════════════════════════════════════════════════════════════════════════
# SHEET 5: Stress Testing / Scenario Analysis
# ══════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Stress Testing")
title_banner(ws5, "STRESS TESTING & SCENARIO ANALYSIS", end_col="E", tab_color="8E44AD")

ws5.merge_cells("A2:E2")
ws5["A2"] = "Estimated fund impact under various stress scenarios. Based on historical factor sensitivities and hypothetical shocks."
ws5["A2"].font = Font(name="Arial", italic=True, size=9, color=DARK_BLUE)
ws5["A2"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)

ws5["A4"] = "HISTORICAL SCENARIOS"
ws5["A4"].font = Font(name="Arial", bold=True, size=12, color=DARK_BLUE)

row_h = 5
stress_headers = ["Scenario", "Estimated Impact", "Severity", "Description"]
for c, h in enumerate(stress_headers, 1):
    ws5.cell(row=row_h, column=c, value=h)
style_header(ws5, row_h, len(stress_headers))

descriptions = {
    "2008 GFC (Sep-Nov)": "Global equity crash, credit freeze. Trend-followers generally profited from short equities.",
    "2020 COVID Crash (Mar)": "Rapid equity selloff. Mixed for CTAs — too fast for some trend signals.",
    "2022 Rate Shock (H1)": "Rising rates, falling bonds. Strong period for trend-following (short bonds worked).",
    "Equity -20% Shock": "Hypothetical equity crash. AHL's low beta limits direct equity losses.",
    "Bond -10% Shock": "Hypothetical bond selloff. Impact depends on current bond positioning.",
    "Commodity +30% Spike": "Hypothetical commodity rally. Positive if AHL is long commodities.",
    "USD +10% Rally": "Strong dollar scenario. Modest positive impact from FX trend positions.",
    "Vol Spike (VIX to 40)": "Volatility expansion. Generally positive for trend-followers.",
}

for i, (scenario, impact) in enumerate(SCENARIOS.items()):
    r = row_h + 1 + i
    ws5.cell(row=r, column=1, value=scenario).font = label_font
    cell = ws5.cell(row=r, column=2, value=round(impact, 4))
    cell.number_format = pct_fmt
    if impact >= 0:
        cell.font = Font(name="Arial", bold=True, size=10, color=GREEN)
    else:
        cell.font = Font(name="Arial", bold=True, size=10, color=RED)

    # Severity
    if abs(impact) >= 0.08:
        ws5.cell(row=r, column=3, value="HIGH")
        ws5.cell(row=r, column=3).font = Font(name="Arial", bold=True, color=RED, size=10)
    elif abs(impact) >= 0.04:
        ws5.cell(row=r, column=3, value="MEDIUM")
        ws5.cell(row=r, column=3).font = Font(name="Arial", bold=True, color=AMBER, size=10)
    else:
        ws5.cell(row=r, column=3, value="LOW")
        ws5.cell(row=r, column=3).font = Font(name="Arial", bold=True, color=GREEN, size=10)

    ws5.cell(row=r, column=4, value=descriptions.get(scenario, "")).font = Font(
        name="Arial", size=9, italic=True)
    style_row(ws5, r, len(stress_headers), alt=(i % 2 == 0))

# Scenario bar chart
sc_row = row_h + len(SCENARIOS) + 3
bar2 = BarChart()
bar2.type = "col"
bar2.title = "Estimated Impact by Scenario"
bar2.y_axis.numFmt = '0%'
bar2.style = 10
bar2.width = 30
bar2.height = 16

data_ref = Reference(ws5, min_col=2, min_row=row_h, max_row=row_h + len(SCENARIOS))
cats_ref = Reference(ws5, min_col=1, min_row=row_h + 1, max_row=row_h + len(SCENARIOS))
bar2.add_data(data_ref, titles_from_data=True)
bar2.set_categories(cats_ref)
bar2.series[0].graphicalProperties.solidFill = MED_BLUE
bar2.legend = None
ws5.add_chart(bar2, f"A{sc_row}")

set_widths(ws5, len(stress_headers), default=18)
ws5.column_dimensions["A"].width = 28
ws5.column_dimensions["D"].width = 55


# ══════════════════════════════════════════════════════════════════════════
# SHEET 6: Limit Monitoring (traffic light system)
# ══════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Limit Monitoring")
title_banner(ws6, "RISK LIMIT MONITORING", end_col="F", tab_color=RED)

ws6.merge_cells("A2:F2")
ws6["A2"] = "Traffic light system: GREEN = within limits, AMBER = approaching limit (>80%), RED = limit breached"
ws6["A2"].font = Font(name="Arial", italic=True, size=9, color=DARK_BLUE)
ws6["A2"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)

row_h = 4
limit_headers = ["Risk Metric", "Current Value", "Limit", "Utilisation", "Headroom", "Status"]
for c, h in enumerate(limit_headers, 1):
    ws6.cell(row=row_h, column=c, value=h)
style_header(ws6, row_h, len(limit_headers))

limit_checks = [
    ("Max Drawdown", current_dd, LIMITS["Max Drawdown Limit"], True),
    ("Monthly VaR (95%)", var_95, LIMITS["Monthly VaR Limit (95%)"], True),
    ("Annualised Volatility", ann_vol, LIMITS["Annualised Volatility Limit"], False),
    ("Beta vs Equity", beta, LIMITS["Beta Limit (vs Equity)"], False),
    ("Max Single Asset Weight", max(e["weight"] for e in EXPOSURES.values()),
     LIMITS["Concentration Limit (single asset)"], False),
]

for i, (name, current, limit, is_negative) in enumerate(limit_checks):
    r = row_h + 1 + i
    ws6.cell(row=r, column=1, value=name).font = label_font

    fmt = pct_fmt if "VaR" in name or "Drawdown" in name or "Vol" in name or "Weight" in name else ratio_fmt
    ws6.cell(row=r, column=2, value=round(current, 4)).number_format = fmt
    ws6.cell(row=r, column=3, value=round(limit, 4)).number_format = fmt

    # Utilisation (how much of the limit is used)
    if is_negative:
        util = current / limit if limit != 0 else 0  # both negative, so ratio is positive when within
    else:
        util = current / limit if limit != 0 else 0
    ws6.cell(row=r, column=4, value=round(abs(util), 4)).number_format = pct_fmt

    # Headroom
    if is_negative:
        headroom = current - limit  # both negative: positive headroom = within limit
    else:
        headroom = limit - current  # positive: positive headroom = within limit
    ws6.cell(row=r, column=5, value=round(headroom, 4)).number_format = fmt

    # Status
    util_pct = abs(util)
    if util_pct < 0.80:
        traffic_light(ws6, r, 6, "GREEN")
    elif util_pct < 1.0:
        traffic_light(ws6, r, 6, "AMBER")
    else:
        traffic_light(ws6, r, 6, "RED")

    style_row(ws6, r, 5, alt=(i % 2 == 0))

set_widths(ws6, len(limit_headers), default=18)
ws6.column_dimensions["A"].width = 28


# ══════════════════════════════════════════════════════════════════════════
# SHEET 7: Risk Commentary (template for analyst notes)
# ══════════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("Risk Commentary")
title_banner(ws7, f"RISK COMMENTARY — {report_date.upper()}", end_col="B", tab_color="95A5A6")

commentary_sections = [
    ("MARKET ENVIRONMENT", 
     "Describe the key market themes this month (e.g., equity rally, rate changes, "
     "geopolitical events, commodity moves). How did these affect the fund's positioning?"),
    ("FUND PERFORMANCE REVIEW",
     f"The fund returned {latest_month_ret:.1%} this month, bringing YTD performance to {ytd_ret:.1%}. "
     f"Current drawdown stands at {current_dd:.1%} from the peak NAV of ${peak_nav:,.0f}.\n\n"
     "[Analyst: Add context on which positions contributed/detracted]"),
    ("RISK METRICS COMMENTARY",
     f"Annualised volatility is {ann_vol:.1%}, [above/below] the long-term average. "
     f"The Sharpe ratio stands at {sharpe:.2f}. VaR (95%, 1M) is {var_95:.1%}.\n\n"
     "[Analyst: Note any significant changes vs prior month and explain drivers]"),
    ("EXPOSURE COMMENTARY",
     "The fund maintains [long/short/neutral] positioning across major asset classes. "
     "Notable changes this month include:\n"
     "- [Asset class]: [position change and rationale]\n"
     "- [Asset class]: [position change and rationale]"),
    ("LIMIT STATUS",
     "All risk limits are currently [within bounds / approaching limits / breached]. "
     "[If AMBER or RED: explain the situation, expected trajectory, and remediation plan]"),
    ("STRESS TEST OBSERVATIONS",
     "Under the current stress scenarios, the fund's largest potential loss would be "
     "from a [scenario name] event. This is [within / approaching / beyond] acceptable levels.\n\n"
     "[Analyst: Note any scenarios that have become more likely given current market conditions]"),
    ("OUTLOOK & ACTION ITEMS",
     "Key risks to monitor next month:\n"
     "1. [Risk factor]\n"
     "2. [Risk factor]\n"
     "3. [Risk factor]\n\n"
     "Recommended actions:\n"
     "- [Action item]\n"
     "- [Action item]"),
]

r = 3
for title, template in commentary_sections:
    ws7.cell(row=r, column=1, value=title).font = Font(
        name="Arial", bold=True, size=12, color=DARK_BLUE)
    r += 1
    ws7.cell(row=r, column=1, value=template).font = Font(
        name="Arial", size=10, color="666666")
    ws7.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws7.row_dimensions[r].height = 80
    r += 2

ws7.column_dimensions["A"].width = 90
ws7.column_dimensions["B"].width = 20


# ── Save ──────────────────────────────────────────────────────────────────
OUTPUT = "Fund_Risk_Report.xlsx"
wb.save(OUTPUT)
print(f"✅ Risk report saved: {OUTPUT}")
print(f"   Sheets: {wb.sheetnames}")
print(f"   Report date: {report_date}")
print(f"   Fund: {FUND_NAME}")
print(f"   Metrics: NAV=${current_nav:,.0f}, DD={current_dd:.1%}, Vol={ann_vol:.1%}, Sharpe={sharpe:.2f}")
