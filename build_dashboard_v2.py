"""
Fund Performance Attribution & Risk Dashboard - Part 1 (Formula-Driven)
========================================================================
Generates a self-contained Excel workbook where ALL metrics are calculated
via Excel formulas. Just update the Data sheet and everything recalculates.

This script automatically picks up 'returns_df' from the notebook if you've
already loaded data in Cell 3. Otherwise it generates simulated data.
"""

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.utils import get_column_letter

# ── Configuration ──────────────────────────────────────────────────────────
RISK_FREE_RATE = 0.045
START_DATE = "2019-01-01"
END_DATE = "2025-12-31"
BENCHMARK_NAME = "Global Equity Index"


def generate_simulated_data():
    """Fallback: generates realistic simulated data if no real data is loaded."""
    FUNDS = {
        "Fund Alpha": {"annual_ret": 0.09, "annual_vol": 0.14},
        "Peer 2 (Trend-Following)": {"annual_ret": 0.065, "annual_vol": 0.11},
        "Peer 1 (Systematic)": {"annual_ret": 0.075, "annual_vol": 0.125},
    }
    np.random.seed(42)
    dates = pd.date_range(START_DATE, END_DATE, freq="ME")
    n = len(dates)
    all_params = {**FUNDS, BENCHMARK_NAME: {"annual_ret": 0.08, "annual_vol": 0.16}}
    data = {}
    market_factor = np.random.normal(0, 0.04, n)
    for name, params in all_params.items():
        monthly_ret = params["annual_ret"] / 12
        monthly_vol = params["annual_vol"] / np.sqrt(12)
        beta = 0.3 if name != BENCHMARK_NAME else 1.0
        idio = np.random.normal(monthly_ret * (1 - beta), monthly_vol * 0.7, n)
        returns = beta * market_factor + idio
        for i, d in enumerate(dates):
            if d.year == 2020 and d.month == 3:
                returns[i] -= np.random.uniform(0.04, 0.10)
            if d.year == 2022 and d.month in [6, 9]:
                returns[i] -= np.random.uniform(0.02, 0.05)
        data[name] = returns
    return pd.DataFrame(data, index=dates)


# ── Pick up data from notebook OR generate simulated ──────────────────────
try:
    # If returns_df already exists (loaded in notebook Cell 3), use it
    _ = returns_df  # noqa: F821
    # Make sure benchmark is the last column (convention used throughout)
    if BENCHMARK_NAME in returns_df.columns:
        cols = [c for c in returns_df.columns if c != BENCHMARK_NAME] + [BENCHMARK_NAME]
        returns_df = returns_df[cols]
    print(f"✅ Using REAL data from notebook ({len(returns_df)} months)")
    print(f"   Funds: {list(returns_df.columns)}")
except NameError:
    returns_df = generate_simulated_data()
    print(f"⚠️  No data loaded in notebook — using SIMULATED data ({len(returns_df)} months)")

FUND_NAMES = list(returns_df.columns)
NUM_FUNDS = len(FUND_NAMES)
dates = returns_df.index
N = len(dates)

# ── Styles ─────────────────────────────────────────────────────────────────
DARK_BLUE = "1B2A4A"
MED_BLUE = "2E5090"
LIGHT_BLUE = "D6E4F0"
WHITE = "FFFFFF"
LIGHT_GREY = "F2F2F2"
GREEN = "27AE60"
RED = "E74C3C"
CHART_COLOURS = ["2E5090", "27AE60", "E67E22", "E74C3C"]

hdr_font = Font(name="Arial", bold=True, color=WHITE, size=11)
hdr_fill = PatternFill("solid", fgColor=DARK_BLUE)
hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
sub_font = Font(name="Arial", bold=True, color=DARK_BLUE, size=10)
sub_fill = PatternFill("solid", fgColor=LIGHT_BLUE)
data_font = Font(name="Arial", size=10)
label_font = Font(name="Arial", bold=True, size=10, color=DARK_BLUE)
thin_border = Border(bottom=Side(style="thin", color="D0D0D0"))


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cl = ws.cell(row=row, column=c)
        cl.font, cl.fill, cl.alignment = hdr_font, hdr_fill, hdr_align


def style_row(ws, row, cols, alt=False):
    fill = PatternFill("solid", fgColor=LIGHT_GREY) if alt else PatternFill("solid", fgColor=WHITE)
    for c in range(1, cols + 1):
        cl = ws.cell(row=row, column=c)
        cl.font = data_font
        cl.fill = fill
        cl.border = thin_border
        cl.alignment = Alignment(horizontal="center" if c > 1 else "left")


def set_widths(ws, cols, default=15):
    for c in range(1, cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = default
    ws.column_dimensions["A"].width = 28


def title_banner(ws, text, merge_end="F1", tab_color=DARK_BLUE):
    ws.merge_cells(f"A1:{merge_end}")
    ws["A1"] = text
    ws["A1"].font = Font(name="Arial", bold=True, color=WHITE, size=14)
    ws["A1"].fill = PatternFill("solid", fgColor=DARK_BLUE)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 38
    ws.sheet_properties.tabColor = tab_color


wb = Workbook()

# Column letters for fund data (B=fund1, C=fund2, D=fund3, E=benchmark)
FUND_COLS = [get_column_letter(c) for c in range(2, 2 + NUM_FUNDS)]
DATA_ROW_START = 3  # first data row in Data sheet
DATA_ROW_END = DATA_ROW_START + N - 1

# Helper: reference a fund's return column in the Data sheet
def dr(col_letter, row=None):
    if row:
        return f"Data!{col_letter}{row}"
    return f"Data!{col_letter}{DATA_ROW_START}:{col_letter}{DATA_ROW_END}"

def dr_abs(col_letter):
    return f"Data!{col_letter}${DATA_ROW_START}:{col_letter}${DATA_ROW_END}"


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 1: Data (raw monthly returns — the input sheet)
# ═══════════════════════════════════════════════════════════════════════════
ws_data = wb.active
ws_data.title = "Data"
ws_data.sheet_properties.tabColor = "95A5A6"

ws_data["A1"] = "MONTHLY RETURN DATA — PASTE YOUR DATA HERE"
ws_data["A1"].font = Font(name="Arial", bold=True, color=DARK_BLUE, size=12)
ws_data.merge_cells(f"A1:{get_column_letter(NUM_FUNDS+1)}1")

# Headers
row_h = 2
ws_data.cell(row=row_h, column=1, value="Date")
for j, name in enumerate(FUND_NAMES):
    ws_data.cell(row=row_h, column=j + 2, value=name)
style_header(ws_data, row_h, NUM_FUNDS + 1)

# Data
for i, (dt, vals) in enumerate(returns_df.iterrows()):
    r = DATA_ROW_START + i
    ws_data.cell(row=r, column=1, value=dt).number_format = "MMM-YY"
    for j, v in enumerate(vals):
        ws_data.cell(row=r, column=j + 2, value=round(v, 6)).number_format = '0.0000%'

set_widths(ws_data, NUM_FUNDS + 1)

# Config cells (below data) for risk-free rate so formulas can reference it
CONFIG_ROW = DATA_ROW_END + 3
ws_data.cell(row=CONFIG_ROW, column=1, value="Risk-Free Rate (annual)").font = label_font
ws_data.cell(row=CONFIG_ROW, column=2, value=RISK_FREE_RATE).number_format = '0.0%'
ws_data.cell(row=CONFIG_ROW, column=2).font = Font(name="Arial", size=10, color="0000FF")

ws_data.cell(row=CONFIG_ROW + 1, column=1, value="Number of Months").font = label_font
ws_data.cell(row=CONFIG_ROW + 1, column=2).value = f'=COUNTA(Data!A{DATA_ROW_START}:A{DATA_ROW_END})'

RF_CELL = f"Data!$B${CONFIG_ROW}"  # absolute ref to risk-free rate
N_CELL = f"Data!$B${CONFIG_ROW + 1}"  # absolute ref to month count


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 2: Summary Dashboard (ALL formulas)
# ═══════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Summary Dashboard")
title_banner(ws, "FUND PERFORMANCE ATTRIBUTION & RISK DASHBOARD",
             merge_end=f"{get_column_letter(NUM_FUNDS+1)}1", tab_color=DARK_BLUE)

ws.merge_cells(f"A2:{get_column_letter(NUM_FUNDS+1)}2")
ws["A2"] = f"All metrics are FORMULA-DRIVEN from the Data sheet  |  Update returns in Data sheet → metrics auto-recalculate"
ws["A2"].font = Font(name="Arial", italic=True, color=DARK_BLUE, size=9)
ws["A2"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)
ws["A2"].alignment = Alignment(horizontal="center")

# Metric table
ROW_START = 4
headers = ["Metric"] + FUND_NAMES
for c, h in enumerate(headers, 1):
    ws.cell(row=ROW_START, column=c, value=h)
style_header(ws, ROW_START, len(headers))

# Build formulas for each metric × each fund
# fc = fund column letter in Data sheet, idx = column index in this sheet
metrics = []

def add_metric(name, formulas, fmt='0.0%'):
    """formulas: list of Excel formula strings, one per fund."""
    metrics.append((name, formulas, fmt))

for fi, fc in enumerate(FUND_COLS):
    pass  # we'll build per-metric below

# We'll build formulas per fund column
def fund_range(fc):
    return f"Data!{fc}${DATA_ROW_START}:{fc}${DATA_ROW_END}"

def bench_range():
    return fund_range(FUND_COLS[-1])  # last column is benchmark

# ── Metric definitions as Excel formulas ──
metric_defs = []

# 1) Annualised Return = (PRODUCT(1+ri))^(12/n) - 1
#    Excel doesn't have PRODUCT over (1+range), so we use EXP(SUMPRODUCT(LN(1+range)))
def ann_return_formula(fc):
    rng = fund_range(fc)
    return f'=EXP(SUMPRODUCT(LN(1+{rng}))*(12/{N_CELL}))-1'
metric_defs.append(("Annualised Return", ann_return_formula, '0.0%'))

# 2) Annualised Volatility = STDEV(range) * SQRT(12)
def ann_vol_formula(fc):
    return f'=STDEV({fund_range(fc)})*SQRT(12)'
metric_defs.append(("Annualised Volatility", ann_vol_formula, '0.0%'))

# 3) Sharpe Ratio = (Ann Return - Rf) / Ann Vol
# We'll reference the cells we just created (rows 5 and 6 in this sheet)
# Row 5 = Ann Return, Row 6 = Ann Vol
def sharpe_formula(fc, col_idx):
    col_l = get_column_letter(col_idx)
    return f'=({col_l}{ROW_START+1}-{RF_CELL})/{col_l}{ROW_START+2}'
metric_defs.append(("Sharpe Ratio", sharpe_formula, '0.00'))

# 4) Sortino Ratio = (Ann Return - Rf) / (Downside Dev * SQRT(12))
#    Downside dev: SQRT(SUMPRODUCT((range<0)*range^2)/COUNT(range))
def sortino_formula(fc, col_idx):
    rng = fund_range(fc)
    col_l = get_column_letter(col_idx)
    downside = f'SQRT(SUMPRODUCT(({rng}<0)*{rng}^2)/COUNTIF({rng},"<0"))*SQRT(12)'
    return f'=({col_l}{ROW_START+1}-{RF_CELL})/({downside})'
metric_defs.append(("Sortino Ratio", sortino_formula, '0.00'))

# 5) Max Drawdown — requires cumulative product which is hard in a single formula
#    We'll compute it via a helper column approach on a hidden Calc sheet
# For now, use a simplified approximation via the Calc sheet (built below)
# Placeholder — will be filled after Calc sheet
metric_defs.append(("Max Drawdown", "CALC_SHEET", '0.0%'))

# 6) Calmar Ratio = Ann Return / |Max Drawdown|
metric_defs.append(("Calmar Ratio", "CALC_SHEET", '0.00'))

# 7) VaR (95%, 1-month) = PERCENTILE(range, 0.05)
def var_formula(fc):
    return f'=PERCENTILE({fund_range(fc)},0.05)'
metric_defs.append(("VaR (95%, 1-month)", var_formula, '0.0%'))

# 8) CVaR (95%, 1-month) = AVERAGEIF(range, "<="&VaR)
def cvar_formula(fc, col_idx):
    rng = fund_range(fc)
    col_l = get_column_letter(col_idx)
    var_row = ROW_START + 7  # VaR is metric index 7 (0-based), row = ROW_START+1+6
    return f'=AVERAGEIF({rng},"<="&{col_l}{var_row})'
metric_defs.append(("CVaR (95%, 1-month)", cvar_formula, '0.0%'))

# 9) Win Rate = COUNTIF(>0) / COUNT
def winrate_formula(fc):
    rng = fund_range(fc)
    return f'=COUNTIF({rng},">0")/COUNT({rng})'
metric_defs.append(("Win Rate", winrate_formula, '0.0%'))

# 10) Best Month
def best_formula(fc):
    return f'=MAX({fund_range(fc)})'
metric_defs.append(("Best Month", best_formula, '0.0%'))

# 11) Worst Month
def worst_formula(fc):
    return f'=MIN({fund_range(fc)})'
metric_defs.append(("Worst Month", worst_formula, '0.0%'))

# 12) Beta = COVARIANCE.S(fund,bench) / VAR.S(bench)
def beta_formula(fc):
    if fc == FUND_COLS[-1]:  # benchmark vs itself
        return 1.0
    # Use SUMPRODUCT for covariance/variance (works in both Excel and LibreOffice)
    f_rng = fund_range(fc)
    b_rng = bench_range()
    n = N_CELL
    # Cov = (SUMPRODUCT(fund,bench) - SUM(fund)*SUM(bench)/n) / (n-1)
    # Var = (SUMPRODUCT(bench,bench) - SUM(bench)^2/n) / (n-1)
    # Beta = Cov/Var, the (n-1) cancels out
    cov = f'SUMPRODUCT({f_rng},{b_rng})-SUMPRODUCT({f_rng})*SUMPRODUCT({b_rng})/{n}'
    var = f'SUMPRODUCT({b_rng},{b_rng})-SUMPRODUCT({b_rng})^2/{n}'
    return f'=({cov})/({var})'
    # Note: COVARIANCE.S and VAR.S must be uppercase for LibreOffice
metric_defs.append(("Beta vs Benchmark", beta_formula, '0.00'))

# 13) Alpha = Ann.Ret_fund - (Rf + Beta * (Ann.Ret_bench - Rf))
def alpha_formula(fc, col_idx):
    if fc == FUND_COLS[-1]:
        return 0.0
    col_l = get_column_letter(col_idx)
    bench_col_l = get_column_letter(NUM_FUNDS + 1)  # benchmark is last
    ann_ret_row = ROW_START + 1  # row 5
    beta_row = ROW_START + 1 + 12  # Beta is metric index 12 → row 5+12=17... 
    # Actually we need to count: metric_defs indices:
    # 0=AnnRet(r5), 1=AnnVol(r6), 2=Sharpe(r7), 3=Sortino(r8), 4=MaxDD(r9),
    # 5=Calmar(r10), 6=VaR(r11), 7=CVaR(r12), 8=WinRate(r13), 9=Best(r14),
    # 10=Worst(r15), 11=Beta(r16), 12=Alpha(r17)
    # So Beta is at ROW_START+1+11 = row 16
    beta_row = ROW_START + 1 + 11  # row 16
    return f'={col_l}{ann_ret_row}-({RF_CELL}+{col_l}{beta_row}*({bench_col_l}{ann_ret_row}-{RF_CELL}))'
metric_defs.append(("Alpha vs Benchmark", alpha_formula, '0.0%'))

# 14) Tracking Error = STDEV(fund-bench)*SQRT(12) — needs Calc sheet difference column
metric_defs.append(("Tracking Error", "CALC_SHEET", '0.0%'))

# 15) Information Ratio = (Mean(fund-bench)*12) / Tracking Error
metric_defs.append(("Information Ratio", "CALC_SHEET", '0.00'))

# Track which metrics need special benchmark handling (benchmark vs itself = 0)
BENCHMARK_ZERO_METRICS = {"Tracking Error", "Information Ratio"}


# ═══════════════════════════════════════════════════════════════════════════
# CALC SHEET (helper columns for drawdown, tracking error, etc.)
# ═══════════════════════════════════════════════════════════════════════════
ws_calc = wb.create_sheet("Calc")
ws_calc.sheet_properties.tabColor = "95A5A6"
ws_calc["A1"] = "HELPER CALCULATIONS (cumulative products, drawdowns, excess returns)"
ws_calc["A1"].font = Font(name="Arial", bold=True, size=10, color=DARK_BLUE)

# Layout: For each fund, we need columns for:
#   Cumulative Product, Running Peak, Drawdown, Excess Return (vs bench)
# That's 4 cols per fund + 1 date column
calc_headers = ["Date"]
calc_col_map = {}  # fund_index -> {"cum": col, "peak": col, "dd": col, "excess": col}
col = 2
for fi, fname in enumerate(FUND_NAMES):
    calc_col_map[fi] = {
        "cum": get_column_letter(col),
        "peak": get_column_letter(col + 1),
        "dd": get_column_letter(col + 2),
        "excess": get_column_letter(col + 3),
    }
    calc_headers.extend([
        f"{fname} Cum", f"{fname} Peak", f"{fname} DD", f"{fname} Excess"
    ])
    col += 4

# Write headers
for c, h in enumerate(calc_headers, 1):
    ws_calc.cell(row=2, column=c, value=h)
    ws_calc.cell(row=2, column=c).font = Font(name="Arial", bold=True, size=8)

# Write date references and formulas
CALC_DATA_START = 3
for i in range(N):
    r = CALC_DATA_START + i
    ws_calc.cell(row=r, column=1).value = f'=Data!A{DATA_ROW_START + i}'
    ws_calc.cell(row=r, column=1).number_format = "MMM-YY"

    for fi in range(NUM_FUNDS):
        fc = FUND_COLS[fi]  # Data sheet column for this fund
        cm = calc_col_map[fi]
        cum_c = cm["cum"]
        peak_c = cm["peak"]
        dd_c = cm["dd"]
        excess_c = cm["excess"]
        bench_fc = FUND_COLS[-1]  # benchmark column in Data

        data_cell = f"Data!{fc}{DATA_ROW_START + i}"

        # Cumulative product: row 1 = (1+r1), row n = prev*(1+rn)
        if i == 0:
            ws_calc[f"{cum_c}{r}"] = f"=1+{data_cell}"
        else:
            ws_calc[f"{cum_c}{r}"] = f"={cum_c}{r-1}*(1+{data_cell})"

        # Running peak = MAX of this cum and previous peak
        if i == 0:
            ws_calc[f"{peak_c}{r}"] = f"={cum_c}{r}"
        else:
            ws_calc[f"{peak_c}{r}"] = f"=MAX({cum_c}{r},{peak_c}{r-1})"

        # Drawdown = (cum - peak) / peak
        ws_calc[f"{dd_c}{r}"] = f"=({cum_c}{r}-{peak_c}{r})/{peak_c}{r}"
        ws_calc[f"{dd_c}{r}"].number_format = '0.00%'

        # Excess return vs benchmark
        bench_cell = f"Data!{bench_fc}{DATA_ROW_START + i}"
        ws_calc[f"{excess_c}{r}"] = f"={data_cell}-{bench_cell}"

CALC_DATA_END = CALC_DATA_START + N - 1

# Summary row: Max Drawdown, Tracking Error, Info Ratio for each fund
CALC_SUMMARY_ROW = CALC_DATA_END + 2
ws_calc.cell(row=CALC_SUMMARY_ROW, column=1, value="Max Drawdown").font = label_font
ws_calc.cell(row=CALC_SUMMARY_ROW + 1, column=1, value="Tracking Error").font = label_font
ws_calc.cell(row=CALC_SUMMARY_ROW + 2, column=1, value="Mean Excess (monthly)").font = label_font
ws_calc.cell(row=CALC_SUMMARY_ROW + 3, column=1, value="Info Ratio").font = label_font

for fi in range(NUM_FUNDS):
    cm = calc_col_map[fi]
    dd_range = f"{cm['dd']}{CALC_DATA_START}:{cm['dd']}{CALC_DATA_END}"
    excess_range = f"{cm['excess']}{CALC_DATA_START}:{cm['excess']}{CALC_DATA_END}"

    # Use the cum column for output placement
    out_col_idx = 2 + fi * 4  # first col of this fund's block
    out_col = get_column_letter(out_col_idx)

    # Max Drawdown = MIN(drawdown range)
    ws_calc[f"{out_col}{CALC_SUMMARY_ROW}"] = f"=MIN({dd_range})"
    ws_calc[f"{out_col}{CALC_SUMMARY_ROW}"].number_format = '0.00%'

    # Tracking Error = STDEV(excess)*SQRT(12)
    ws_calc[f"{out_col}{CALC_SUMMARY_ROW+1}"] = f"=STDEV({excess_range})*SQRT(12)"
    ws_calc[f"{out_col}{CALC_SUMMARY_ROW+1}"].number_format = '0.00%'

    # Mean Excess (monthly)
    ws_calc[f"{out_col}{CALC_SUMMARY_ROW+2}"] = f"=AVERAGE({excess_range})"
    ws_calc[f"{out_col}{CALC_SUMMARY_ROW+2}"].number_format = '0.0000%'

    # Info Ratio = Mean Excess * 12 / Tracking Error (with IFERROR for benchmark)
    ws_calc[f"{out_col}{CALC_SUMMARY_ROW+3}"] = f"=IFERROR({out_col}{CALC_SUMMARY_ROW+2}*12/{out_col}{CALC_SUMMARY_ROW+1},0)"
    ws_calc[f"{out_col}{CALC_SUMMARY_ROW+3}"].number_format = '0.00'

# Auto-size calc columns
for c in range(1, len(calc_headers) + 1):
    ws_calc.column_dimensions[get_column_letter(c)].width = 13


# ═══════════════════════════════════════════════════════════════════════════
# NOW: Write all formulas into Summary Dashboard
# ═══════════════════════════════════════════════════════════════════════════
for mi, (metric_name, formula_fn, fmt) in enumerate(metric_defs):
    r = ROW_START + 1 + mi
    ws.cell(row=r, column=1, value=metric_name).font = label_font

    for fi, fc in enumerate(FUND_COLS):
        col_idx = fi + 2  # column in summary sheet
        cell = ws.cell(row=r, column=col_idx)

        if formula_fn == "CALC_SHEET":
            # For benchmark column, Tracking Error and Info Ratio = 0
            if fi == NUM_FUNDS - 1 and metric_name in BENCHMARK_ZERO_METRICS:
                cell.value = 0
            else:
                # Reference Calc sheet summary
                out_col_idx_calc = 2 + fi * 4
                out_col_calc = get_column_letter(out_col_idx_calc)

                if metric_name == "Max Drawdown":
                    cell.value = f"=Calc!{out_col_calc}{CALC_SUMMARY_ROW}"
                elif metric_name == "Calmar Ratio":
                    ann_ret_cell = f"{get_column_letter(col_idx)}{ROW_START+1}"
                    dd_cell = f"{get_column_letter(col_idx)}{r-1}"
                    cell.value = f'={ann_ret_cell}/ABS({dd_cell})'
                elif metric_name == "Tracking Error":
                    cell.value = f"=Calc!{out_col_calc}{CALC_SUMMARY_ROW+1}"
                elif metric_name == "Information Ratio":
                    cell.value = f"=Calc!{out_col_calc}{CALC_SUMMARY_ROW+3}"
        elif callable(formula_fn):
            import inspect
            sig = inspect.signature(formula_fn)
            if len(sig.parameters) == 1:
                cell.value = formula_fn(fc)
            elif len(sig.parameters) == 2:
                cell.value = formula_fn(fc, col_idx)
        else:
            cell.value = formula_fn

        cell.number_format = fmt

    style_row(ws, r, len(headers), alt=(mi % 2 == 0))

set_widths(ws, len(headers))


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 3: Calendar Year Returns (formulas)
# ═══════════════════════════════════════════════════════════════════════════
ws_yr = wb.create_sheet("Calendar Year Returns")
title_banner(ws_yr, "CALENDAR YEAR RETURNS", merge_end=f"{get_column_letter(NUM_FUNDS+1)}1", tab_color=MED_BLUE)

years = sorted(returns_df.index.year.unique())

row_h = 3
headers_yr = ["Year"] + FUND_NAMES
for c, h in enumerate(headers_yr, 1):
    ws_yr.cell(row=row_h, column=c, value=h)
style_header(ws_yr, row_h, len(headers_yr))

for yi, year in enumerate(years):
    r = row_h + 1 + yi
    ws_yr.cell(row=r, column=1, value=year).number_format = "0"

    # Find row range in Data sheet for this year
    year_mask = returns_df.index.year == year
    first_idx = returns_df.index.get_loc(returns_df.index[year_mask][0])
    last_idx = returns_df.index.get_loc(returns_df.index[year_mask][-1])
    data_r_start = DATA_ROW_START + first_idx
    data_r_end = DATA_ROW_START + last_idx

    for fi, fc in enumerate(FUND_COLS):
        # Calendar year return = PRODUCT(1+ri) - 1 via EXP(SUM(LN(1+range)))-1
        rng = f"Data!{fc}{data_r_start}:{fc}{data_r_end}"
        cell = ws_yr.cell(row=r, column=fi + 2)
        cell.value = f"=EXP(SUMPRODUCT(LN(1+{rng})))-1"
        cell.number_format = '0.0%'

    style_row(ws_yr, r, len(headers_yr), alt=(yi % 2 == 0))

set_widths(ws_yr, len(headers_yr))

# Bar chart
n_yrs = len(years)
bar_chart = BarChart()
bar_chart.type = "col"
bar_chart.title = "Calendar Year Returns"
bar_chart.y_axis.numFmt = '0%'
bar_chart.style = 10
bar_chart.width = 30
bar_chart.height = 14

for j in range(NUM_FUNDS):
    ref = Reference(ws_yr, min_col=j + 2, min_row=row_h, max_row=row_h + n_yrs)
    bar_chart.add_data(ref, titles_from_data=True)
    bar_chart.series[j].graphicalProperties.solidFill = CHART_COLOURS[j % len(CHART_COLOURS)]

cats = Reference(ws_yr, min_col=1, min_row=row_h + 1, max_row=row_h + n_yrs)
bar_chart.set_categories(cats)
bar_chart.legend.position = "b"
ws_yr.add_chart(bar_chart, f"A{row_h + n_yrs + 3}")


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 4: Monthly Heatmap for Fund Alpha (formulas referencing Data sheet)
# ═══════════════════════════════════════════════════════════════════════════
ws_mth = wb.create_sheet("Monthly Returns - AHL")
title_banner(ws_mth, "MAN AHL — MONTHLY RETURN HEATMAP", merge_end="N1", tab_color=GREEN)

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

row_h = 3
mth_headers = ["Year"] + MONTHS + ["YTD"]
for c, h in enumerate(mth_headers, 1):
    ws_mth.cell(row=row_h, column=c, value=h)
style_header(ws_mth, row_h, len(mth_headers))

ahl_col = FUND_COLS[0]  # Fund Alpha is first fund

for yi, year in enumerate(years):
    r = row_h + 1 + yi
    ws_mth.cell(row=r, column=1, value=year).number_format = "0"

    year_mask = returns_df.index.year == year
    year_months = returns_df.index[year_mask]

    for dt in year_months:
        month_num = dt.month
        data_row = DATA_ROW_START + returns_df.index.get_loc(dt)
        cell = ws_mth.cell(row=r, column=month_num + 1)
        cell.value = f"=Data!{ahl_col}{data_row}"
        cell.number_format = '0.0%'
        cell.alignment = Alignment(horizontal="center")

    # YTD formula
    first_idx = returns_df.index.get_loc(year_months[0])
    last_idx = returns_df.index.get_loc(year_months[-1])
    ytd_rng = f"Data!{ahl_col}{DATA_ROW_START + first_idx}:{ahl_col}{DATA_ROW_START + last_idx}"
    ytd_cell = ws_mth.cell(row=r, column=14)
    ytd_cell.value = f"=EXP(SUMPRODUCT(LN(1+{ytd_rng})))-1"
    ytd_cell.number_format = '0.0%'
    ytd_cell.font = Font(name="Arial", bold=True, size=10)

for c in range(1, len(mth_headers) + 1):
    ws_mth.column_dimensions[get_column_letter(c)].width = 9
ws_mth.column_dimensions["A"].width = 10
ws_mth.column_dimensions["N"].width = 10


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 5: Drawdown Analysis (references Calc sheet)
# ═══════════════════════════════════════════════════════════════════════════
ws_dd = wb.create_sheet("Drawdown Analysis")
title_banner(ws_dd, "DRAWDOWN ANALYSIS (UNDERWATER CHART)", merge_end=f"{get_column_letter(NUM_FUNDS+1)}1", tab_color=RED)

row_h = 3
dd_headers = ["Date"] + FUND_NAMES
for c, h in enumerate(dd_headers, 1):
    ws_dd.cell(row=row_h, column=c, value=h)
style_header(ws_dd, row_h, len(dd_headers))

for i in range(N):
    r = row_h + 1 + i
    ws_dd.cell(row=r, column=1).value = f"=Calc!A{CALC_DATA_START + i}"
    ws_dd.cell(row=r, column=1).number_format = "MMM-YY"
    for fi in range(NUM_FUNDS):
        dd_col = calc_col_map[fi]["dd"]
        cell = ws_dd.cell(row=r, column=fi + 2)
        cell.value = f"=Calc!{dd_col}{CALC_DATA_START + i}"
        cell.number_format = '0.0%'

# Drawdown chart
dd_chart = LineChart()
dd_chart.title = "Underwater Chart (Drawdowns)"
dd_chart.y_axis.title = "Drawdown"
dd_chart.y_axis.numFmt = '0%'
dd_chart.style = 10
dd_chart.width = 32
dd_chart.height = 14

for j in range(NUM_FUNDS):
    ref = Reference(ws_dd, min_col=j + 2, min_row=row_h, max_row=row_h + N)
    dd_chart.add_data(ref, titles_from_data=True)
    dd_chart.series[j].graphicalProperties.line.width = 18000
    dd_chart.series[j].graphicalProperties.line.solidFill = CHART_COLOURS[j % len(CHART_COLOURS)]

cats = Reference(ws_dd, min_col=1, min_row=row_h + 1, max_row=row_h + N)
dd_chart.set_categories(cats)
dd_chart.legend.position = "b"
ws_dd.add_chart(dd_chart, f"A{row_h + N + 3}")
set_widths(ws_dd, len(dd_headers))


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 6: Cumulative Growth (formulas from Calc sheet)
# ═══════════════════════════════════════════════════════════════════════════
ws_cum = wb.create_sheet("Cumulative Growth")
title_banner(ws_cum, "CUMULATIVE GROWTH OF $100", merge_end=f"{get_column_letter(NUM_FUNDS+1)}1", tab_color=MED_BLUE)

row_h = 3
cum_headers = ["Date"] + FUND_NAMES
for c, h in enumerate(cum_headers, 1):
    ws_cum.cell(row=row_h, column=c, value=h)
style_header(ws_cum, row_h, len(cum_headers))

for i in range(N):
    r = row_h + 1 + i
    ws_cum.cell(row=r, column=1).value = f"=Calc!A{CALC_DATA_START + i}"
    ws_cum.cell(row=r, column=1).number_format = "MMM-YY"
    for fi in range(NUM_FUNDS):
        cum_col = calc_col_map[fi]["cum"]
        cell = ws_cum.cell(row=r, column=fi + 2)
        cell.value = f"=Calc!{cum_col}{CALC_DATA_START + i}*100"
        cell.number_format = '$#,##0.00'

# Growth chart
growth_chart = LineChart()
growth_chart.title = "Cumulative Growth of $100"
growth_chart.y_axis.title = "Value ($)"
growth_chart.style = 10
growth_chart.width = 32
growth_chart.height = 16

for j in range(NUM_FUNDS):
    ref = Reference(ws_cum, min_col=j + 2, min_row=row_h, max_row=row_h + N)
    growth_chart.add_data(ref, titles_from_data=True)
    growth_chart.series[j].graphicalProperties.line.width = 22000
    growth_chart.series[j].graphicalProperties.line.solidFill = CHART_COLOURS[j % len(CHART_COLOURS)]

cats = Reference(ws_cum, min_col=1, min_row=row_h + 1, max_row=row_h + N)
growth_chart.set_categories(cats)
growth_chart.legend.position = "b"
ws_cum.add_chart(growth_chart, f"A{row_h + N + 3}")
set_widths(ws_cum, len(cum_headers))


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 7: Methodology
# ═══════════════════════════════════════════════════════════════════════════
ws_meth = wb.create_sheet("Methodology")
title_banner(ws_meth, "METHODOLOGY & DEFINITIONS", merge_end="C1", tab_color="95A5A6")

definitions = [
    ("Metric", "Formula", "Description"),
    ("Annualised Return", "EXP(SUMPRODUCT(LN(1+r)) × 12/n) - 1", "Geometric mean annual return"),
    ("Annualised Volatility", "STDEV(monthly) × √12", "Standard deviation of returns, annualised"),
    ("Sharpe Ratio", "(Ann. Return - Rf) / Ann. Vol", "Return per unit of total risk"),
    ("Sortino Ratio", "(Ann. Return - Rf) / Downside Dev", "Return per unit of downside risk only"),
    ("Max Drawdown", "MIN((Cum - Peak) / Peak)", "Largest peak-to-trough decline"),
    ("Calmar Ratio", "Ann. Return / |Max Drawdown|", "Return per unit of max drawdown"),
    ("VaR (95%)", "PERCENTILE(returns, 5%)", "Worst expected monthly loss at 95% confidence"),
    ("CVaR (95%)", "AVERAGEIF(returns ≤ VaR)", "Average loss in worst 5% of months"),
    ("Beta", "COVARIANCE(fund, bench) / VAR(bench)", "Sensitivity to benchmark"),
    ("Alpha (CAPM)", "Rp - [Rf + β(Rm - Rf)]", "Excess return above CAPM prediction"),
    ("Tracking Error", "STDEV(fund - bench) × √12", "Volatility of active returns"),
    ("Information Ratio", "Mean(excess) × 12 / TE", "Active return per unit of active risk"),
    ("Win Rate", "Count(r > 0) / Count(r)", "Percentage of positive months"),
]

row = 3
for c, h in enumerate(definitions[0], 1):
    ws_meth.cell(row=row, column=c, value=h)
style_header(ws_meth, row, 3)

for i, (m, f, d) in enumerate(definitions[1:], 1):
    r = row + i
    ws_meth.cell(row=r, column=1, value=m).font = label_font
    ws_meth.cell(row=r, column=2, value=f).font = Font(name="Consolas", size=10)
    ws_meth.cell(row=r, column=3, value=d).font = data_font
    style_row(ws_meth, r, 3, alt=(i % 2 == 0))

ws_meth.column_dimensions["A"].width = 24
ws_meth.column_dimensions["B"].width = 38
ws_meth.column_dimensions["C"].width = 45

note_r = row + len(definitions) + 1
ws_meth["A" + str(note_r)] = "HOW TO UPDATE WITH REAL DATA"
ws_meth["A" + str(note_r)].font = Font(name="Arial", bold=True, size=11, color=DARK_BLUE)
instructions = [
    "1. Open the Jupyter Notebook (fund_performance_notebook.ipynb)",
    "2. Run the cells to fetch live data from Yahoo Finance",
    "3. The script overwrites the Data sheet with real returns",
    "4. All formulas across all sheets auto-recalculate",
    "5. Tickers: AHLPX (Fund Alpha), AQMIX (Peer 1), URTH (Global Equity Index)",
    f"6. Risk-free rate is in cell B{CONFIG_ROW} of the Data sheet (currently {RISK_FREE_RATE:.1%})",
]
for i, txt in enumerate(instructions):
    ws_meth[f"A{note_r + 1 + i}"] = txt
    ws_meth[f"A{note_r + 1 + i}"].font = Font(name="Consolas", size=10)


# ═══════════════════════════════════════════════════════════════════════════
# Reorder sheets: Data last (or keep as reference tab)
# ═══════════════════════════════════════════════════════════════════════════
desired_order = ["Summary Dashboard", "Calendar Year Returns", "Monthly Returns - AHL",
                 "Drawdown Analysis", "Cumulative Growth", "Methodology", "Data", "Calc"]
sheet_map = {s: i for i, s in enumerate(desired_order)}
wb._sheets.sort(key=lambda s: sheet_map.get(s.title, 99))

# ── Save ───────────────────────────────────────────────────────────────────
OUTPUT = "Fund_Performance_Dashboard_v2.xlsx"
wb.save(OUTPUT)
print(f"✅ Workbook saved: {OUTPUT}")
print(f"   Sheets: {wb.sheetnames}")
print(f"   Data rows: {N} months ({START_DATE} to {END_DATE})")
print(f"   Formula-driven: Yes — update Data sheet and everything recalculates")
