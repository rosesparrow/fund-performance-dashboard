"""
Part 2: Performance Attribution Analysis
==========================================
Two attribution models in one workbook:

PART A — Classic Brinson Attribution (Equity Sector ETFs)
  Uses a simulated equity portfolio vs S&P 500 benchmark to demonstrate
  the standard Brinson-Fachler model: Allocation + Selection + Interaction.
  This is the textbook model that interview questions are based on.

PART B — Factor Attribution for Fund Alpha
  Decomposes AHL's returns into contributions from asset class factors:
  Equities, Bonds, Commodities, and FX. This is more realistic for a
  managed futures fund that trades across multiple asset classes.

OUTPUT: Formula-driven Excel workbook.
"""

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

# ── Styles (matching Part 1) ──────────────────────────────────────────────
DARK_BLUE = "1B2A4A"
MED_BLUE = "2E5090"
LIGHT_BLUE = "D6E4F0"
WHITE = "FFFFFF"
LIGHT_GREY = "F2F2F2"
GREEN = "27AE60"
RED = "E74C3C"
ORANGE = "E67E22"
CHART_COLOURS = ["2E5090", "27AE60", "E67E22", "E74C3C", "8E44AD", "1ABC9C", "F39C12",
                 "2C3E50", "D35400", "7F8C8D", "C0392B"]

hdr_font = Font(name="Arial", bold=True, color=WHITE, size=11)
hdr_fill = PatternFill("solid", fgColor=DARK_BLUE)
hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
data_font = Font(name="Arial", size=10)
label_font = Font(name="Arial", bold=True, size=10, color=DARK_BLUE)
pct_fmt = '0.0%'
pct_fmt2 = '0.00%'
thin_border = Border(bottom=Side(style="thin", color="D0D0D0"))


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cl = ws.cell(row=row, column=c)
        cl.font, cl.fill, cl.alignment = hdr_font, hdr_fill, hdr_align


def style_row(ws, row, cols, alt=False):
    fill = PatternFill("solid", fgColor=LIGHT_GREY) if alt else PatternFill("solid", fgColor=WHITE)
    for c in range(1, cols + 1):
        cl = ws.cell(row=row, column=c)
        cl.font = data_font
        cl.fill = fill
        cl.border = thin_border
        cl.alignment = Alignment(horizontal="center" if c > 1 else "left")


def title_banner(ws, text, end_col="H", tab_color=DARK_BLUE, row_num=1):
    ws.merge_cells(f"A{row_num}:{end_col}{row_num}")
    ws[f"A{row_num}"] = text
    ws[f"A{row_num}"].font = Font(name="Arial", bold=True, color=WHITE, size=14)
    ws[f"A{row_num}"].fill = PatternFill("solid", fgColor=DARK_BLUE)
    ws[f"A{row_num}"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row_num].height = 38
    ws.sheet_properties.tabColor = tab_color


def set_widths(ws, cols, default=15):
    for c in range(1, cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = default
    ws.column_dimensions["A"].width = 24


# ══════════════════════════════════════════════════════════════════════════
# PART A: BRINSON ATTRIBUTION DATA
# ══════════════════════════════════════════════════════════════════════════
# We create a realistic equity portfolio with sector weights and returns
# that differ from the benchmark, so the attribution effects are clear.

SECTORS = [
    "Technology", "Healthcare", "Financials", "Consumer Disc.",
    "Industrials", "Energy", "Consumer Staples", "Utilities",
    "Materials", "Real Estate", "Communication Svcs"
]
NUM_SECTORS = len(SECTORS)

# Simulated data for one quarter (this would come from portfolio systems)
np.random.seed(123)

# Benchmark weights (roughly S&P 500-like)
bench_weights = np.array([0.28, 0.13, 0.13, 0.10, 0.09, 0.05, 0.07, 0.03, 0.03, 0.03, 0.06])
bench_weights = bench_weights / bench_weights.sum()  # ensure sums to 1

# Portfolio weights (our fund's active bets)
# Overweight Tech & Healthcare, underweight Energy & Utilities
port_weights = np.array([0.35, 0.16, 0.10, 0.08, 0.08, 0.02, 0.06, 0.01, 0.04, 0.04, 0.06])
port_weights = port_weights / port_weights.sum()

# Benchmark sector returns (one quarter)
bench_returns = np.array([0.08, 0.04, 0.06, 0.03, 0.05, -0.02, 0.02, 0.01, 0.03, -0.01, 0.04])

# Portfolio sector returns (our stock picks within each sector)
# Better picks in Tech and Healthcare, worse in Financials
port_returns = np.array([0.11, 0.06, 0.04, 0.04, 0.05, -0.03, 0.03, 0.01, 0.02, 0.00, 0.05])

# Total benchmark return
total_bench_return = (bench_weights * bench_returns).sum()
total_port_return = (port_weights * port_returns).sum()
total_excess = total_port_return - total_bench_return


# ══════════════════════════════════════════════════════════════════════════
# PART B: AHL FACTOR ATTRIBUTION DATA
# ══════════════════════════════════════════════════════════════════════════
# For a managed futures fund, we break down returns by asset class factor.
# This is a simplified version of what Fund Alpha actually does.

FACTORS = ["Equities", "Fixed Income", "Commodities", "FX", "Residual/Alpha"]
NUM_FACTORS = len(FACTORS)

# Simulated quarterly data (4 quarters)
QUARTERS = ["Q1 2024", "Q2 2024", "Q3 2024", "Q4 2024"]
NUM_QUARTERS = len(QUARTERS)

np.random.seed(456)
# Factor contributions (what each asset class contributed to AHL's return)
# Rows = quarters, Cols = factors
factor_contributions = np.array([
    [0.015, 0.008, 0.012, -0.003, 0.005],   # Q1: equities & commodities strong
    [-0.010, 0.015, 0.005, 0.008, 0.003],    # Q2: bonds rally, equities weak
    [0.020, -0.005, 0.018, 0.006, 0.004],    # Q3: equities & commodities strong
    [-0.015, 0.020, -0.008, 0.010, 0.006],   # Q4: bonds strong, equities & commodities weak
])

# Total AHL return per quarter = sum of factor contributions
ahl_quarterly_returns = factor_contributions.sum(axis=1)


# ══════════════════════════════════════════════════════════════════════════
# BUILD THE WORKBOOK
# ══════════════════════════════════════════════════════════════════════════
wb = Workbook()

# ──────────────────────────────────────────────────────────────────────────
# SHEET 1: Brinson Overview (explanation + summary)
# ──────────────────────────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Brinson Overview"
title_banner(ws1, "BRINSON ATTRIBUTION — HOW IT WORKS", end_col="H", tab_color=DARK_BLUE)

# Explanation section
explanations = [
    ("", ""),
    ("WHAT IS BRINSON ATTRIBUTION?", ""),
    ("", "Brinson attribution answers: WHY did the portfolio outperform or underperform the benchmark?"),
    ("", "It breaks the excess return into three components:"),
    ("", ""),
    ("1. ALLOCATION EFFECT", "Did we overweight sectors that performed well?"),
    ("   Formula:", "(Portfolio Weight - Benchmark Weight) × (Benchmark Sector Return - Total Benchmark Return)"),
    ("   Example:", "We had 35% in Tech vs benchmark's 28%. Tech returned 8%. Benchmark total was 5%."),
    ("   ", "Allocation = (0.35 - 0.28) × (0.08 - 0.05) = +0.21% contribution"),
    ("", ""),
    ("2. SELECTION EFFECT", "Did we pick better stocks within each sector?"),
    ("   Formula:", "Benchmark Weight × (Portfolio Sector Return - Benchmark Sector Return)"),
    ("   Example:", "Benchmark had 28% in Tech. Our Tech stocks returned 11% vs benchmark Tech's 8%."),
    ("   ", "Selection = 0.28 × (0.11 - 0.08) = +0.84% contribution"),
    ("", ""),
    ("3. INTERACTION EFFECT", "Did we overweight sectors where we also picked well?"),
    ("   Formula:", "(Portfolio Weight - Benchmark Weight) × (Portfolio Sector Return - Benchmark Sector Return)"),
    ("   Example:", "We were +7% overweight Tech AND our Tech picks beat by +3%."),
    ("   ", "Interaction = 0.07 × 0.03 = +0.21% contribution"),
    ("", ""),
    ("KEY INSIGHT:", "Allocation + Selection + Interaction = Total Excess Return (always, exactly)"),
    ("", "This is what makes Brinson powerful — it decomposes perfectly with no residual."),
]

for i, (label, desc) in enumerate(explanations):
    r = 3 + i
    ws1.cell(row=r, column=1, value=label)
    ws1.cell(row=r, column=2, value=desc)
    if label and label[0] != " " and label != "":
        ws1.cell(row=r, column=1).font = Font(name="Arial", bold=True, size=10, color=DARK_BLUE)
        ws1.cell(row=r, column=2).font = Font(name="Arial", bold=True, size=10)
    elif "Formula:" in label:
        ws1.cell(row=r, column=1).font = Font(name="Arial", italic=True, size=10)
        ws1.cell(row=r, column=2).font = Font(name="Consolas", size=10)
    else:
        ws1.cell(row=r, column=1).font = data_font
        ws1.cell(row=r, column=2).font = data_font

ws1.column_dimensions["A"].width = 28
ws1.column_dimensions["B"].width = 85


# ──────────────────────────────────────────────────────────────────────────
# SHEET 2: Sector Data (input — weights and returns)
# ──────────────────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Sector Data")
title_banner(ws2, "SECTOR WEIGHTS & RETURNS — INPUT DATA", end_col="F", tab_color="95A5A6")

ws2.merge_cells("A2:F2")
ws2["A2"] = "This sheet contains the raw inputs. Update these to analyse different periods. All attribution formulas reference this sheet."
ws2["A2"].font = Font(name="Arial", italic=True, size=9, color=DARK_BLUE)
ws2["A2"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)

row = 4
headers = ["Sector", "Portfolio Weight", "Benchmark Weight", "Portfolio Return", "Benchmark Return", "Active Weight"]
for c, h in enumerate(headers, 1):
    ws2.cell(row=row, column=c, value=h)
style_header(ws2, row, len(headers))

DATA_START = row + 1
for i, sector in enumerate(SECTORS):
    r = DATA_START + i
    ws2.cell(row=r, column=1, value=sector).font = label_font
    ws2.cell(row=r, column=2, value=round(port_weights[i], 4)).number_format = pct_fmt
    ws2.cell(row=r, column=3, value=round(bench_weights[i], 4)).number_format = pct_fmt
    ws2.cell(row=r, column=4, value=round(port_returns[i], 4)).number_format = pct_fmt
    ws2.cell(row=r, column=5, value=round(bench_returns[i], 4)).number_format = pct_fmt
    # Active Weight = Portfolio Weight - Benchmark Weight (FORMULA)
    ws2.cell(row=r, column=6).value = f"=B{r}-C{r}"
    ws2.cell(row=r, column=6).number_format = pct_fmt
    # Blue font for inputs (industry standard)
    for c in [2, 3, 4, 5]:
        ws2.cell(row=r, column=c).font = Font(name="Arial", size=10, color="0000FF")
    style_row(ws2, r, len(headers), alt=(i % 2 == 0))

DATA_END = DATA_START + NUM_SECTORS - 1

# Totals row
TOTAL_ROW = DATA_END + 1
ws2.cell(row=TOTAL_ROW, column=1, value="TOTAL").font = Font(name="Arial", bold=True, size=10)
ws2.cell(row=TOTAL_ROW, column=2).value = f"=SUM(B{DATA_START}:B{DATA_END})"
ws2.cell(row=TOTAL_ROW, column=2).number_format = pct_fmt
ws2.cell(row=TOTAL_ROW, column=3).value = f"=SUM(C{DATA_START}:C{DATA_END})"
ws2.cell(row=TOTAL_ROW, column=3).number_format = pct_fmt
# Weighted returns (total portfolio and benchmark return)
ws2.cell(row=TOTAL_ROW, column=4).value = f"=SUMPRODUCT(B{DATA_START}:B{DATA_END},D{DATA_START}:D{DATA_END})"
ws2.cell(row=TOTAL_ROW, column=4).number_format = pct_fmt
ws2.cell(row=TOTAL_ROW, column=5).value = f"=SUMPRODUCT(C{DATA_START}:C{DATA_END},E{DATA_START}:E{DATA_END})"
ws2.cell(row=TOTAL_ROW, column=5).number_format = pct_fmt
ws2.cell(row=TOTAL_ROW, column=6).value = f"=B{TOTAL_ROW}-C{TOTAL_ROW}"
ws2.cell(row=TOTAL_ROW, column=6).number_format = pct_fmt
for c in range(1, len(headers) + 1):
    ws2.cell(row=TOTAL_ROW, column=c).font = Font(name="Arial", bold=True, size=10)
    ws2.cell(row=TOTAL_ROW, column=c).border = Border(
        top=Side(style="double", color="000000"),
        bottom=Side(style="double", color="000000"))

# Reference cells for total returns
TOTAL_PORT_RET_CELL = f"D{TOTAL_ROW}"   # total portfolio return
TOTAL_BENCH_RET_CELL = f"E{TOTAL_ROW}"  # total benchmark return

set_widths(ws2, len(headers), default=18)
ws2.column_dimensions["A"].width = 22


# ──────────────────────────────────────────────────────────────────────────
# SHEET 3: Brinson Attribution (all formulas)
# ──────────────────────────────────────────────────────────────────────────
ws3 = wb.create_sheet("Brinson Attribution")
title_banner(ws3, "BRINSON-FACHLER ATTRIBUTION ANALYSIS", end_col="H", tab_color=MED_BLUE)

ws3.merge_cells("A2:H2")
ws3["A2"] = "All cells are FORMULAS referencing the Sector Data sheet. Change the inputs and this recalculates."
ws3["A2"].font = Font(name="Arial", italic=True, size=9, color=DARK_BLUE)
ws3["A2"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)

# Summary box at top
ws3["A4"] = "SUMMARY"
ws3["A4"].font = Font(name="Arial", bold=True, size=12, color=DARK_BLUE)

summary_items = [
    ("Portfolio Return:", f"='Sector Data'!{TOTAL_PORT_RET_CELL}", pct_fmt2),
    ("Benchmark Return:", f"='Sector Data'!{TOTAL_BENCH_RET_CELL}", pct_fmt2),
    ("Excess Return:", f"='Sector Data'!{TOTAL_PORT_RET_CELL}-'Sector Data'!{TOTAL_BENCH_RET_CELL}", pct_fmt2),
]
for i, (label, formula, fmt) in enumerate(summary_items):
    r = 5 + i
    ws3.cell(row=r, column=1, value=label).font = label_font
    ws3.cell(row=r, column=2).value = formula
    ws3.cell(row=r, column=2).number_format = fmt
    ws3.cell(row=r, column=2).font = Font(name="Arial", bold=True, size=11)

# Attribution decomposition summary
ws3["A9"] = "ATTRIBUTION DECOMPOSITION"
ws3["A9"].font = Font(name="Arial", bold=True, size=12, color=DARK_BLUE)

# These will sum the detail rows (built below)
DETAIL_START = 15
DETAIL_END = DETAIL_START + NUM_SECTORS - 1

decomp_items = [
    ("Allocation Effect:", f"=SUM(E{DETAIL_START}:E{DETAIL_END})", GREEN),
    ("Selection Effect:", f"=SUM(F{DETAIL_START}:F{DETAIL_END})", MED_BLUE),
    ("Interaction Effect:", f"=SUM(G{DETAIL_START}:G{DETAIL_END})", ORANGE),
    ("Total (should = Excess):", f"=SUM(H{DETAIL_START}:H{DETAIL_END})", DARK_BLUE),
]
for i, (label, formula, color) in enumerate(decomp_items):
    r = 10 + i
    ws3.cell(row=r, column=1, value=label).font = Font(name="Arial", bold=True, size=10, color=color)
    ws3.cell(row=r, column=2).value = formula
    ws3.cell(row=r, column=2).number_format = pct_fmt2
    ws3.cell(row=r, column=2).font = Font(name="Arial", bold=True, size=11, color=color)

# Detail table
row_h = DETAIL_START - 1
detail_headers = ["Sector", "Active Weight", "Portfolio Return", "Benchmark Return",
                  "Allocation", "Selection", "Interaction", "Total Contribution"]
for c, h in enumerate(detail_headers, 1):
    ws3.cell(row=row_h, column=c, value=h)
style_header(ws3, row_h, len(detail_headers))

for i in range(NUM_SECTORS):
    r = DETAIL_START + i
    sd_r = DATA_START + i  # corresponding row in Sector Data sheet

    # Sector name
    ws3.cell(row=r, column=1).value = f"='Sector Data'!A{sd_r}"
    ws3.cell(row=r, column=1).font = label_font

    # Active Weight = Port Weight - Bench Weight
    ws3.cell(row=r, column=2).value = f"='Sector Data'!F{sd_r}"
    ws3.cell(row=r, column=2).number_format = pct_fmt2

    # Portfolio Return (reference)
    ws3.cell(row=r, column=3).value = f"='Sector Data'!D{sd_r}"
    ws3.cell(row=r, column=3).number_format = pct_fmt

    # Benchmark Return (reference)
    ws3.cell(row=r, column=4).value = f"='Sector Data'!E{sd_r}"
    ws3.cell(row=r, column=4).number_format = pct_fmt

    # ── THE BRINSON FORMULAS ──

    # ALLOCATION EFFECT:
    # (Port Weight - Bench Weight) × (Bench Sector Return - Total Bench Return)
    # "We were overweight this sector, and the sector beat/lagged the overall benchmark"
    ws3.cell(row=r, column=5).value = (
        f"=('Sector Data'!B{sd_r}-'Sector Data'!C{sd_r})"
        f"*('Sector Data'!E{sd_r}-'Sector Data'!{TOTAL_BENCH_RET_CELL})"
    )
    ws3.cell(row=r, column=5).number_format = pct_fmt2

    # SELECTION EFFECT:
    # Bench Weight × (Port Sector Return - Bench Sector Return)
    # "Within this sector, our picks beat/lagged the benchmark's stocks"
    ws3.cell(row=r, column=6).value = (
        f"='Sector Data'!C{sd_r}"
        f"*('Sector Data'!D{sd_r}-'Sector Data'!E{sd_r})"
    )
    ws3.cell(row=r, column=6).number_format = pct_fmt2

    # INTERACTION EFFECT:
    # (Port Weight - Bench Weight) × (Port Sector Return - Bench Sector Return)
    # "We overweighted AND picked well (or underweighted AND picked poorly)"
    ws3.cell(row=r, column=7).value = (
        f"=('Sector Data'!B{sd_r}-'Sector Data'!C{sd_r})"
        f"*('Sector Data'!D{sd_r}-'Sector Data'!E{sd_r})"
    )
    ws3.cell(row=r, column=7).number_format = pct_fmt2

    # TOTAL = Allocation + Selection + Interaction
    ws3.cell(row=r, column=8).value = f"=E{r}+F{r}+G{r}"
    ws3.cell(row=r, column=8).number_format = pct_fmt2

    style_row(ws3, r, len(detail_headers), alt=(i % 2 == 0))

# Totals row
tr = DETAIL_END + 1
ws3.cell(row=tr, column=1, value="TOTAL").font = Font(name="Arial", bold=True, size=10)
for c in range(5, 9):
    col_l = get_column_letter(c)
    ws3.cell(row=tr, column=c).value = f"=SUM({col_l}{DETAIL_START}:{col_l}{DETAIL_END})"
    ws3.cell(row=tr, column=c).number_format = pct_fmt2
    ws3.cell(row=tr, column=c).font = Font(name="Arial", bold=True, size=10)
    ws3.cell(row=tr, column=c).border = Border(
        top=Side(style="double", color="000000"),
        bottom=Side(style="double", color="000000"))

set_widths(ws3, len(detail_headers), default=16)
ws3.column_dimensions["A"].width = 22

# ── Waterfall Chart (Allocation | Selection | Interaction) ──
chart_row = tr + 3
ws3.cell(row=chart_row, column=1, value="ATTRIBUTION WATERFALL").font = Font(
    name="Arial", bold=True, size=12, color=DARK_BLUE)

# Prepare chart data in a small table
wr = chart_row + 1
ws3.cell(row=wr, column=1, value="Component")
ws3.cell(row=wr, column=2, value="Contribution")
style_header(ws3, wr, 2)

chart_items = [
    ("Allocation", f"=SUM(E{DETAIL_START}:E{DETAIL_END})"),
    ("Selection", f"=SUM(F{DETAIL_START}:F{DETAIL_END})"),
    ("Interaction", f"=SUM(G{DETAIL_START}:G{DETAIL_END})"),
    ("Total Excess", f"=SUM(H{DETAIL_START}:H{DETAIL_END})"),
]
for i, (label, formula) in enumerate(chart_items):
    r = wr + 1 + i
    ws3.cell(row=r, column=1, value=label).font = label_font
    ws3.cell(row=r, column=2).value = formula
    ws3.cell(row=r, column=2).number_format = pct_fmt2

# Bar chart
bar = BarChart()
bar.type = "col"
bar.title = "Attribution Breakdown"
bar.y_axis.numFmt = '0.00%'
bar.y_axis.title = "Contribution to Excess Return"
bar.style = 10
bar.width = 22
bar.height = 14

data_ref = Reference(ws3, min_col=2, min_row=wr, max_row=wr + len(chart_items))
cats_ref = Reference(ws3, min_col=1, min_row=wr + 1, max_row=wr + len(chart_items))
bar.add_data(data_ref, titles_from_data=True)
bar.set_categories(cats_ref)
bar.series[0].graphicalProperties.solidFill = MED_BLUE
bar.legend = None

ws3.add_chart(bar, f"D{chart_row}")

# ── Sector contribution bar chart ──
sector_chart_row = chart_row + 18
ws3.cell(row=sector_chart_row, column=1, value="SECTOR TOTAL CONTRIBUTIONS").font = Font(
    name="Arial", bold=True, size=12, color=DARK_BLUE)

sector_bar = BarChart()
sector_bar.type = "col"
sector_bar.title = "Total Contribution by Sector"
sector_bar.y_axis.numFmt = '0.00%'
sector_bar.style = 10
sector_bar.width = 30
sector_bar.height = 14

data_ref = Reference(ws3, min_col=8, min_row=row_h, max_row=DETAIL_END)
cats_ref = Reference(ws3, min_col=1, min_row=DETAIL_START, max_row=DETAIL_END)
sector_bar.add_data(data_ref, titles_from_data=True)
sector_bar.set_categories(cats_ref)
sector_bar.series[0].graphicalProperties.solidFill = MED_BLUE
sector_bar.legend = None

ws3.add_chart(sector_bar, f"A{sector_chart_row + 1}")


# ──────────────────────────────────────────────────────────────────────────
# SHEET 4: AHL Factor Attribution Data
# ──────────────────────────────────────────────────────────────────────────
ws4 = wb.create_sheet("AHL Factor Data")
title_banner(ws4, "MAN AHL — FACTOR CONTRIBUTION DATA", end_col="G", tab_color="27AE60")

ws4.merge_cells("A2:G2")
ws4["A2"] = "How much did each asset class contribute to AHL's total return each quarter? This is the input data."
ws4["A2"].font = Font(name="Arial", italic=True, size=9, color=DARK_BLUE)
ws4["A2"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)

row = 4
fheaders = ["Quarter"] + FACTORS + ["Total Return"]
for c, h in enumerate(fheaders, 1):
    ws4.cell(row=row, column=c, value=h)
style_header(ws4, row, len(fheaders))

F_DATA_START = row + 1
for i, quarter in enumerate(QUARTERS):
    r = F_DATA_START + i
    ws4.cell(row=r, column=1, value=quarter).font = label_font
    for j in range(NUM_FACTORS):
        cell = ws4.cell(row=r, column=j + 2, value=round(factor_contributions[i, j], 4))
        cell.number_format = pct_fmt2
        cell.font = Font(name="Arial", size=10, color="0000FF")
    # Total Return = sum of factors (FORMULA)
    last_factor_col = get_column_letter(NUM_FACTORS + 1)
    ws4.cell(row=r, column=NUM_FACTORS + 2).value = f"=SUM(B{r}:{last_factor_col}{r})"
    ws4.cell(row=r, column=NUM_FACTORS + 2).number_format = pct_fmt2
    ws4.cell(row=r, column=NUM_FACTORS + 2).font = Font(name="Arial", bold=True, size=10)
    style_row(ws4, r, len(fheaders), alt=(i % 2 == 0))

F_DATA_END = F_DATA_START + NUM_QUARTERS - 1

# Totals / Averages
agg_row = F_DATA_END + 1
ws4.cell(row=agg_row, column=1, value="TOTAL (cumulative)").font = Font(name="Arial", bold=True, size=10)
for c in range(2, len(fheaders) + 1):
    col_l = get_column_letter(c)
    ws4.cell(row=agg_row, column=c).value = f"=SUM({col_l}{F_DATA_START}:{col_l}{F_DATA_END})"
    ws4.cell(row=agg_row, column=c).number_format = pct_fmt2
    ws4.cell(row=agg_row, column=c).font = Font(name="Arial", bold=True, size=10)

avg_row = agg_row + 1
ws4.cell(row=avg_row, column=1, value="AVERAGE (quarterly)").font = Font(name="Arial", bold=True, size=10)
for c in range(2, len(fheaders) + 1):
    col_l = get_column_letter(c)
    ws4.cell(row=avg_row, column=c).value = f"=AVERAGE({col_l}{F_DATA_START}:{col_l}{F_DATA_END})"
    ws4.cell(row=avg_row, column=c).number_format = pct_fmt2

set_widths(ws4, len(fheaders), default=16)
ws4.column_dimensions["A"].width = 24


# ──────────────────────────────────────────────────────────────────────────
# SHEET 5: AHL Factor Attribution Analysis
# ──────────────────────────────────────────────────────────────────────────
ws5 = wb.create_sheet("AHL Factor Attribution")
title_banner(ws5, "MAN AHL — FACTOR ATTRIBUTION ANALYSIS", end_col="G", tab_color="27AE60")

ws5.merge_cells("A2:G2")
ws5["A2"] = "Which asset classes drove AHL's performance? All formulas reference the AHL Factor Data sheet."
ws5["A2"].font = Font(name="Arial", italic=True, size=9, color=DARK_BLUE)
ws5["A2"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)

# Summary
ws5["A4"] = "CUMULATIVE FACTOR CONTRIBUTIONS"
ws5["A4"].font = Font(name="Arial", bold=True, size=12, color=DARK_BLUE)

for i, factor in enumerate(FACTORS):
    r = 5 + i
    ws5.cell(row=r, column=1, value=factor).font = label_font
    col_l = get_column_letter(i + 2)
    ws5.cell(row=r, column=2).value = f"='AHL Factor Data'!{col_l}{agg_row}"
    ws5.cell(row=r, column=2).number_format = pct_fmt2
    ws5.cell(row=r, column=2).font = Font(name="Arial", bold=True, size=11)

r_total = 5 + NUM_FACTORS
ws5.cell(row=r_total, column=1, value="TOTAL RETURN").font = Font(name="Arial", bold=True, size=10, color=DARK_BLUE)
total_col = get_column_letter(NUM_FACTORS + 2)
ws5.cell(row=r_total, column=2).value = f"='AHL Factor Data'!{total_col}{agg_row}"
ws5.cell(row=r_total, column=2).number_format = pct_fmt2
ws5.cell(row=r_total, column=2).font = Font(name="Arial", bold=True, size=11, color=DARK_BLUE)

# Quarterly breakdown table
q_start_row = r_total + 3
ws5.cell(row=q_start_row, column=1, value="QUARTERLY BREAKDOWN").font = Font(
    name="Arial", bold=True, size=12, color=DARK_BLUE)

qr_h = q_start_row + 1
qr_headers = ["Quarter"] + FACTORS + ["Total"]
for c, h in enumerate(qr_headers, 1):
    ws5.cell(row=qr_h, column=c, value=h)
style_header(ws5, qr_h, len(qr_headers))

for i in range(NUM_QUARTERS):
    r = qr_h + 1 + i
    fd_r = F_DATA_START + i
    ws5.cell(row=r, column=1).value = f"='AHL Factor Data'!A{fd_r}"
    ws5.cell(row=r, column=1).font = label_font
    for j in range(NUM_FACTORS):
        col_l = get_column_letter(j + 2)
        ws5.cell(row=r, column=j + 2).value = f"='AHL Factor Data'!{col_l}{fd_r}"
        ws5.cell(row=r, column=j + 2).number_format = pct_fmt2
    total_col = get_column_letter(NUM_FACTORS + 2)
    ws5.cell(row=r, column=NUM_FACTORS + 2).value = f"='AHL Factor Data'!{total_col}{fd_r}"
    ws5.cell(row=r, column=NUM_FACTORS + 2).number_format = pct_fmt2
    ws5.cell(row=r, column=NUM_FACTORS + 2).font = Font(name="Arial", bold=True, size=10)
    style_row(ws5, r, len(qr_headers), alt=(i % 2 == 0))

set_widths(ws5, len(qr_headers), default=16)
ws5.column_dimensions["A"].width = 22

# ── Factor contribution stacked bar chart ──
chart_start = qr_h + NUM_QUARTERS + 3
ws5.cell(row=chart_start, column=1, value="FACTOR CONTRIBUTIONS BY QUARTER").font = Font(
    name="Arial", bold=True, size=12, color=DARK_BLUE)

stacked = BarChart()
stacked.type = "col"
stacked.grouping = "stacked"
stacked.title = "AHL Return Decomposition by Factor"
stacked.y_axis.numFmt = '0.0%'
stacked.y_axis.title = "Return Contribution"
stacked.style = 10
stacked.width = 28
stacked.height = 16

# One series per factor
for j in range(NUM_FACTORS):
    ref = Reference(ws5, min_col=j + 2, min_row=qr_h, max_row=qr_h + NUM_QUARTERS)
    stacked.add_data(ref, titles_from_data=True)
    stacked.series[j].graphicalProperties.solidFill = CHART_COLOURS[j % len(CHART_COLOURS)]

cats_ref = Reference(ws5, min_col=1, min_row=qr_h + 1, max_row=qr_h + NUM_QUARTERS)
stacked.set_categories(cats_ref)
stacked.legend.position = "b"

ws5.add_chart(stacked, f"A{chart_start + 1}")

# ── Cumulative factor pie chart-style bar ──
pie_row = chart_start + 20
ws5.cell(row=pie_row, column=1, value="CUMULATIVE CONTRIBUTION BY FACTOR").font = Font(
    name="Arial", bold=True, size=12, color=DARK_BLUE)

cum_bar = BarChart()
cum_bar.type = "col"
cum_bar.title = "Total Contribution by Asset Class (Full Year)"
cum_bar.y_axis.numFmt = '0.0%'
cum_bar.style = 10
cum_bar.width = 22
cum_bar.height = 14

# Build small reference table for the chart
pr = pie_row + 1
ws5.cell(row=pr, column=1, value="Factor")
ws5.cell(row=pr, column=2, value="Contribution")
style_header(ws5, pr, 2)
for i, factor in enumerate(FACTORS):
    r = pr + 1 + i
    ws5.cell(row=r, column=1, value=factor).font = label_font
    ws5.cell(row=r, column=2).value = f"=B{5 + i}"
    ws5.cell(row=r, column=2).number_format = pct_fmt2

data_ref = Reference(ws5, min_col=2, min_row=pr, max_row=pr + NUM_FACTORS)
cats_ref = Reference(ws5, min_col=1, min_row=pr + 1, max_row=pr + NUM_FACTORS)
cum_bar.add_data(data_ref, titles_from_data=True)
cum_bar.set_categories(cats_ref)

# Color each bar differently
for i in range(NUM_FACTORS):
    cum_bar.series[0].graphicalProperties.solidFill = MED_BLUE
cum_bar.legend = None

ws5.add_chart(cum_bar, f"D{pie_row}")


# ──────────────────────────────────────────────────────────────────────────
# SHEET 6: Methodology
# ──────────────────────────────────────────────────────────────────────────
ws6 = wb.create_sheet("Methodology")
title_banner(ws6, "ATTRIBUTION METHODOLOGY", end_col="C", tab_color="95A5A6")

definitions = [
    ("Model", "Formula", "What It Measures"),
    ("", "", ""),
    ("BRINSON-FACHLER MODEL", "", "(Classic equity attribution)"),
    ("Allocation Effect", "(Wp - Wb) × (Rb_sector - Rb_total)",
     "Value added by sector weight decisions"),
    ("Selection Effect", "Wb × (Rp_sector - Rb_sector)",
     "Value added by stock picking within sectors"),
    ("Interaction Effect", "(Wp - Wb) × (Rp_sector - Rb_sector)",
     "Combined effect of overweighting + picking well"),
    ("Total Excess", "Sum of all three effects",
     "Must equal Portfolio Return minus Benchmark Return"),
    ("", "", ""),
    ("FACTOR ATTRIBUTION", "", "(For managed futures / multi-asset)"),
    ("Factor Contribution", "Beta_factor × Factor_return",
     "How much return came from each asset class"),
    ("Residual / Alpha", "Total Return - Sum(Factor Contributions)",
     "Return not explained by known factors = manager skill"),
    ("", "", ""),
    ("KEY CONCEPTS", "", ""),
    ("Active Weight", "Portfolio Weight - Benchmark Weight",
     "How much you over/underweight each sector"),
    ("Excess Return", "Portfolio Return - Benchmark Return",
     "Total outperformance (positive) or underperformance (negative)"),
    ("Wp, Wb", "Portfolio and Benchmark weights", ""),
    ("Rp, Rb", "Portfolio and Benchmark returns", ""),
]

row = 3
for c, h in enumerate(definitions[0], 1):
    ws6.cell(row=row, column=c, value=h)
style_header(ws6, row, 3)

for i, (m, f, d) in enumerate(definitions[1:], 1):
    r = row + i
    if m and m == m.upper() and len(m) > 3:
        ws6.cell(row=r, column=1, value=m).font = Font(name="Arial", bold=True, size=11, color=DARK_BLUE)
        ws6.cell(row=r, column=3, value=d).font = Font(name="Arial", italic=True, size=10)
    else:
        ws6.cell(row=r, column=1, value=m).font = label_font
        ws6.cell(row=r, column=2, value=f).font = Font(name="Consolas", size=10)
        ws6.cell(row=r, column=3, value=d).font = data_font
    if m:
        style_row(ws6, r, 3, alt=(i % 2 == 0))

ws6.column_dimensions["A"].width = 24
ws6.column_dimensions["B"].width = 42
ws6.column_dimensions["C"].width = 48


# ── Reorder and save ─────────────────────────────────────────────────────
desired_order = ["Brinson Overview", "Sector Data", "Brinson Attribution",
                 "AHL Factor Data", "AHL Factor Attribution", "Methodology"]
sheet_map = {s: i for i, s in enumerate(desired_order)}
wb._sheets.sort(key=lambda s: sheet_map.get(s.title, 99))

OUTPUT = "Fund_Attribution_Analysis.xlsx"
wb.save(OUTPUT)
print(f"✅ Attribution workbook saved: {OUTPUT}")
print(f"   Sheets: {wb.sheetnames}")
print(f"   Brinson sectors: {NUM_SECTORS}")
print(f"   AHL factor quarters: {NUM_QUARTERS}")
