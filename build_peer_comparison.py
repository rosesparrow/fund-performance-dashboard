"""
Part 3: Peer Group / Competitor Analysis
==========================================
Compares Fund Alpha against a peer group of managed futures funds across
all performance and risk metrics. Produces ranking tables, risk/return
scatter plot data, and rolling comparisons.

Picks up returns_df_peer from the notebook if available,
otherwise generates simulated data.

PEER GROUP:
  Fund Alpha (AHLPX) — Primary trend-following fund
  Peer 1 (Systematic) (AQMIX) — Major systematic futures peer
  Peer 2 (Replication) — Largest MF ETF, replicates CTA index
  Peer 3 (Index-Based) — Transparent index-based approach
  Simplify Managed Futures (CTA) — Strong recent performer
  Global Equity Index (URTH) — Equity benchmark for context
"""

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, ScatterChart, Reference, Series
from openpyxl.utils import get_column_letter
import inspect

# ── Configuration ─────────────────────────────────────────────────────────
RISK_FREE_RATE = 0.045
BENCHMARK_NAME = "Global Equity Index"

PEER_TICKERS = {
    "Fund Alpha": "AHLPX",
    "Peer 1 (Systematic)": "AQMIX",
    "Peer 2 (Replication)": "DBMF",
    "Peer 3 (Index-Based)": "KMLM",
    "Peer 4 (Multi-Strategy)": "CTA",
    "Global Equity Index": "URTH",
}

# ── Try to pick up data from notebook ────────────────────────────────────
try:
    _ = returns_df_peer  # noqa: F821
    print(f"✅ Using REAL peer data from notebook ({len(returns_df_peer)} months)")
    print(f"   Funds: {list(returns_df_peer.columns)}")
    df = returns_df_peer.copy()
except NameError:
    print("⚠️  No peer data loaded — using SIMULATED data")
    np.random.seed(999)
    dates = pd.date_range("2020-01-01", "2025-12-31", freq="ME")
    n = len(dates)
    market = np.random.normal(0, 0.04, n)
    params = {
        "Fund Alpha": (0.09, 0.14, 0.3),
        "Peer 1 (Systematic)": (0.075, 0.125, 0.25),
        "Peer 2 (Replication)": (0.07, 0.12, 0.35),
        "Peer 3 (Index-Based)": (0.055, 0.15, 0.2),
        "Peer 4 (Multi-Strategy)": (0.085, 0.13, 0.3),
        "Global Equity Index": (0.08, 0.16, 1.0),
    }
    data = {}
    for name, (ar, av, beta) in params.items():
        mr, mv = ar / 12, av / np.sqrt(12)
        idio = np.random.normal(mr * (1 - beta), mv * 0.7, n)
        rets = beta * market + idio
        for i, d in enumerate(dates):
            if d.year == 2020 and d.month == 3:
                rets[i] -= np.random.uniform(0.03, 0.09)
            if d.year == 2022 and d.month in [6, 9]:
                rets[i] -= np.random.uniform(0.01, 0.04)
        data[name] = rets
    df = pd.DataFrame(data, index=dates)

FUND_NAMES = list(df.columns)
NUM_FUNDS = len(FUND_NAMES)
N = len(df)
dates = df.index


# ── Metric calculations ──────────────────────────────────────────────────
def calc_all_metrics(returns_df, bench_col, rf=RISK_FREE_RATE):
    """Calculate all metrics for all funds. Returns a DataFrame."""
    bench = returns_df[bench_col]
    results = {}
    for col in returns_df.columns:
        r = returns_df[col]
        n = len(r)
        ann_ret = (1 + r).prod() ** (12 / n) - 1
        ann_vol = r.std() * np.sqrt(12)
        downside_vol = r[r < 0].std() * np.sqrt(12)
        cum = (1 + r).cumprod()
        peak = cum.cummax()
        dd = (cum - peak) / peak
        max_dd = dd.min()
        cov_fb = np.cov(r, bench)[0, 1]
        var_b = np.var(bench, ddof=1)
        beta = cov_fb / var_b if var_b != 0 else 0
        b_ann = (1 + bench).prod() ** (12 / len(bench)) - 1
        alpha = ann_ret - (rf + beta * (b_ann - rf))
        excess = r.values - bench.values
        te = np.std(excess, ddof=1) * np.sqrt(12)
        ir = (np.mean(excess) * 12) / te if te != 0 else 0
        results[col] = {
            "Annualised Return": ann_ret,
            "Annualised Volatility": ann_vol,
            "Sharpe Ratio": (ann_ret - rf) / ann_vol if ann_vol != 0 else 0,
            "Sortino Ratio": (ann_ret - rf) / downside_vol if downside_vol != 0 else 0,
            "Max Drawdown": max_dd,
            "Calmar Ratio": ann_ret / abs(max_dd) if max_dd != 0 else 0,
            "VaR (95%, 1M)": np.percentile(r, 5),
            "CVaR (95%, 1M)": r[r <= np.percentile(r, 5)].mean(),
            "Win Rate": (r > 0).mean(),
            "Best Month": r.max(),
            "Worst Month": r.min(),
            "Beta vs Benchmark": beta if col != bench_col else 1.0,
            "Alpha vs Benchmark": alpha if col != bench_col else 0.0,
            "Tracking Error": te if col != bench_col else 0.0,
            "Information Ratio": ir if col != bench_col else 0.0,
        }
    return pd.DataFrame(results)

metrics = calc_all_metrics(df, BENCHMARK_NAME)


# ── Rankings ──────────────────────────────────────────────────────────────
# For each metric, rank funds (1 = best)
# "Best" depends on the metric: higher is better for returns/ratios, lower for vol/drawdown/VaR
higher_is_better = [
    "Annualised Return", "Sharpe Ratio", "Sortino Ratio", "Calmar Ratio",
    "Win Rate", "Best Month", "Alpha vs Benchmark", "Information Ratio"
]
lower_is_better = [
    "Annualised Volatility", "Max Drawdown", "VaR (95%, 1M)",
    "CVaR (95%, 1M)", "Worst Month", "Beta vs Benchmark", "Tracking Error"
]

rankings = pd.DataFrame(index=metrics.index, columns=metrics.columns)
for metric in metrics.index:
    if metric in higher_is_better:
        rankings.loc[metric] = metrics.loc[metric].rank(ascending=False).astype(int)
    elif metric in lower_is_better:
        # For negative metrics like Max DD, "less negative" = better = higher value
        if metric in ["Max Drawdown", "VaR (95%, 1M)", "CVaR (95%, 1M)", "Worst Month"]:
            rankings.loc[metric] = metrics.loc[metric].rank(ascending=False).astype(int)
        else:
            rankings.loc[metric] = metrics.loc[metric].rank(ascending=True).astype(int)
    else:
        rankings.loc[metric] = metrics.loc[metric].rank(ascending=False).astype(int)

# Overall score: average rank across key metrics (lower = better)
key_metrics = ["Annualised Return", "Sharpe Ratio", "Sortino Ratio", "Max Drawdown",
               "Calmar Ratio", "Alpha vs Benchmark", "Information Ratio"]
overall_rank = rankings.loc[key_metrics].astype(float).mean()
overall_rank_sorted = overall_rank.sort_values()

# Rolling Sharpe (12-month)
def rolling_sharpe(returns_df, window=12, rf=RISK_FREE_RATE):
    def sharpe_w(x):
        ann_ret = (1 + x).prod() ** (12 / len(x)) - 1
        vol = x.std() * np.sqrt(12)
        return (ann_ret - rf) / vol if vol != 0 else 0
    return returns_df.rolling(window).apply(sharpe_w, raw=False).dropna()

rolling_sh = rolling_sharpe(df)

# Calendar year returns
yearly = df.groupby(df.index.year).apply(lambda x: (1 + x).prod() - 1)
yearly.index.name = "Year"


# ── Styles ────────────────────────────────────────────────────────────────
DARK_BLUE = "1B2A4A"
MED_BLUE = "2E5090"
LIGHT_BLUE = "D6E4F0"
WHITE = "FFFFFF"
LIGHT_GREY = "F2F2F2"
GREEN = "27AE60"
RED = "E74C3C"
GOLD = "F4D03F"
CHART_COLOURS = ["2E5090", "27AE60", "E67E22", "E74C3C", "8E44AD", "1ABC9C"]

hdr_font = Font(name="Arial", bold=True, color=WHITE, size=11)
hdr_fill = PatternFill("solid", fgColor=DARK_BLUE)
hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
data_font = Font(name="Arial", size=10)
label_font = Font(name="Arial", bold=True, size=10, color=DARK_BLUE)
thin_border = Border(bottom=Side(style="thin", color="D0D0D0"))
pct_fmt = '0.0%'
ratio_fmt = '0.00'

rank_fills = {
    1: PatternFill("solid", fgColor="27AE60"),  # Green = 1st
    2: PatternFill("solid", fgColor="82E0AA"),  # Light green = 2nd
    3: PatternFill("solid", fgColor="F4D03F"),  # Gold = 3rd
}
rank_fonts = {
    1: Font(name="Arial", bold=True, size=10, color=WHITE),
    2: Font(name="Arial", bold=True, size=10),
    3: Font(name="Arial", bold=True, size=10),
}


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cl = ws.cell(row=row, column=c)
        cl.font, cl.fill, cl.alignment = hdr_font, hdr_fill, hdr_align


def style_row(ws, row, cols, alt=False):
    fill = PatternFill("solid", fgColor=LIGHT_GREY) if alt else PatternFill("solid", fgColor=WHITE)
    for c in range(1, cols + 1):
        cl = ws.cell(row=row, column=c)
        cl.font = data_font
        cl.fill = fill
        cl.border = thin_border
        cl.alignment = Alignment(horizontal="center" if c > 1 else "left")


def title_banner(ws, text, end_col="H", tab_color=DARK_BLUE):
    ws.merge_cells(f"A1:{end_col}1")
    ws["A1"] = text
    ws["A1"].font = Font(name="Arial", bold=True, color=WHITE, size=14)
    ws["A1"].fill = PatternFill("solid", fgColor=DARK_BLUE)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 38
    ws.sheet_properties.tabColor = tab_color


def set_widths(ws, cols, default=15):
    for c in range(1, cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = default
    ws.column_dimensions["A"].width = 24


wb = Workbook()


# ══════════════════════════════════════════════════════════════════════════
# SHEET 1: Peer Comparison Table (metrics + colour-coded rankings)
# ══════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Peer Comparison"
end_col = get_column_letter(NUM_FUNDS + 1)
title_banner(ws1, "MANAGED FUTURES PEER GROUP COMPARISON", end_col=end_col, tab_color=DARK_BLUE)

ws1.merge_cells(f"A2:{end_col}2")
period_start = dates[0].strftime('%b %Y')
period_end = dates[-1].strftime('%b %Y')
ws1["A2"] = f"Period: {period_start} to {period_end}  |  Benchmark: Global Equity Index  |  Risk-Free Rate: {RISK_FREE_RATE:.1%}  |  Green = Best, Gold = 3rd"
ws1["A2"].font = Font(name="Arial", italic=True, color=DARK_BLUE, size=9)
ws1["A2"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)
ws1["A2"].alignment = Alignment(horizontal="center")

row_h = 4
headers = ["Metric"] + FUND_NAMES
for c, h in enumerate(headers, 1):
    ws1.cell(row=row_h, column=c, value=h)
style_header(ws1, row_h, len(headers))

metric_formats = {
    "Annualised Return": pct_fmt, "Annualised Volatility": pct_fmt,
    "Sharpe Ratio": ratio_fmt, "Sortino Ratio": ratio_fmt, "Calmar Ratio": ratio_fmt,
    "Max Drawdown": pct_fmt, "VaR (95%, 1M)": pct_fmt, "CVaR (95%, 1M)": pct_fmt,
    "Win Rate": pct_fmt, "Best Month": pct_fmt, "Worst Month": pct_fmt,
    "Beta vs Benchmark": ratio_fmt, "Alpha vs Benchmark": pct_fmt,
    "Tracking Error": pct_fmt, "Information Ratio": ratio_fmt,
}

for mi, metric in enumerate(metrics.index):
    r = row_h + 1 + mi
    ws1.cell(row=r, column=1, value=metric).font = label_font
    for fi, fund in enumerate(FUND_NAMES):
        cell = ws1.cell(row=r, column=fi + 2)
        cell.value = metrics.loc[metric, fund]
        cell.number_format = metric_formats.get(metric, ratio_fmt)
        # Colour-code top 3
        rank = int(rankings.loc[metric, fund])
        if rank in rank_fills:
            cell.fill = rank_fills[rank]
            cell.font = rank_fonts[rank]
    style_row(ws1, r, 1, alt=(mi % 2 == 0))  # only style col A background

set_widths(ws1, len(headers), default=16)

# ── Overall ranking row ──
or_row = row_h + 1 + len(metrics.index) + 1
ws1.cell(row=or_row, column=1, value="OVERALL SCORE (avg rank)").font = Font(
    name="Arial", bold=True, size=11, color=DARK_BLUE)
ws1.cell(row=or_row + 1, column=1, value="(lower = better)").font = Font(
    name="Arial", italic=True, size=9)

for fi, fund in enumerate(FUND_NAMES):
    cell = ws1.cell(row=or_row, column=fi + 2)
    cell.value = round(overall_rank[fund], 1)
    cell.number_format = '0.0'
    cell.font = Font(name="Arial", bold=True, size=11)
    # Highlight the winner
    if fund == overall_rank_sorted.index[0]:
        cell.fill = PatternFill("solid", fgColor=GREEN)
        cell.font = Font(name="Arial", bold=True, size=11, color=WHITE)


# ══════════════════════════════════════════════════════════════════════════
# SHEET 2: Rankings Table
# ══════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Rankings")
title_banner(ws2, "PEER GROUP RANKINGS (1 = BEST)", end_col=end_col, tab_color=MED_BLUE)

row_h = 3
headers_r = ["Metric"] + FUND_NAMES
for c, h in enumerate(headers_r, 1):
    ws2.cell(row=row_h, column=c, value=h)
style_header(ws2, row_h, len(headers_r))

for mi, metric in enumerate(rankings.index):
    r = row_h + 1 + mi
    ws2.cell(row=r, column=1, value=metric).font = label_font
    for fi, fund in enumerate(FUND_NAMES):
        cell = ws2.cell(row=r, column=fi + 2)
        rank_val = int(rankings.loc[metric, fund])
        cell.value = rank_val
        cell.alignment = Alignment(horizontal="center")
        if rank_val in rank_fills:
            cell.fill = rank_fills[rank_val]
            cell.font = rank_fonts[rank_val]
    style_row(ws2, r, 1, alt=(mi % 2 == 0))

set_widths(ws2, len(headers_r), default=16)


# ══════════════════════════════════════════════════════════════════════════
# SHEET 3: Risk/Return Scatter Plot
# ══════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Risk-Return Plot")
title_banner(ws3, "RISK / RETURN SCATTER PLOT", end_col="H", tab_color="27AE60")

ws3["A3"] = "This plot shows each fund's annualised volatility (risk) vs annualised return."
ws3["A3"].font = Font(name="Arial", italic=True, size=9, color=DARK_BLUE)
ws3["A4"] = "Top-left = best (high return, low risk). Bottom-right = worst."
ws3["A4"].font = Font(name="Arial", italic=True, size=9, color=DARK_BLUE)

# Data table for scatter
row_h = 6
ws3.cell(row=row_h, column=1, value="Fund")
ws3.cell(row=row_h, column=2, value="Ann. Volatility (X)")
ws3.cell(row=row_h, column=3, value="Ann. Return (Y)")
ws3.cell(row=row_h, column=4, value="Sharpe Ratio")
style_header(ws3, row_h, 4)

for i, fund in enumerate(FUND_NAMES):
    r = row_h + 1 + i
    ws3.cell(row=r, column=1, value=fund).font = label_font
    ws3.cell(row=r, column=2, value=metrics.loc["Annualised Volatility", fund]).number_format = pct_fmt
    ws3.cell(row=r, column=3, value=metrics.loc["Annualised Return", fund]).number_format = pct_fmt
    ws3.cell(row=r, column=4, value=metrics.loc["Sharpe Ratio", fund]).number_format = ratio_fmt
    style_row(ws3, r, 4, alt=(i % 2 == 0))

# Scatter chart — use a single series with all funds as data points
# openpyxl scatter needs proper ranges, not single cells
scatter = ScatterChart()
scatter.title = "Risk / Return — Managed Futures Peer Group"
scatter.x_axis.title = "Annualised Volatility (Risk)"
scatter.x_axis.numFmt = '0%'
scatter.y_axis.title = "Annualised Return"
scatter.y_axis.numFmt = '0%'
scatter.style = 10
scatter.width = 28
scatter.height = 18

# All funds as one series so they appear as separate points
x_ref = Reference(ws3, min_col=2, min_row=row_h + 1, max_row=row_h + NUM_FUNDS)
y_ref = Reference(ws3, min_col=3, min_row=row_h + 1, max_row=row_h + NUM_FUNDS)
series = Series(y_ref, x_ref, title="Funds")
series.graphicalProperties.line.noFill = True  # no connecting line

# Add data labels showing fund names
from openpyxl.chart.label import DataLabelList
series.dLbls = DataLabelList()
series.dLbls.showCatName = False
series.dLbls.showVal = False
series.dLbls.showSerName = False

scatter.series.append(series)
scatter.legend = None

# Add a text note identifying each point
label_row = row_h + NUM_FUNDS + 2
ws3.cell(row=label_row, column=1, value="Note: Points left-to-right by volatility:").font = Font(
    name="Arial", italic=True, size=9, color=DARK_BLUE)
# Sort funds by volatility for the legend
sorted_funds = sorted(FUND_NAMES, key=lambda f: metrics.loc["Annualised Volatility", f])
fund_labels = ", ".join([f"{f} ({metrics.loc['Annualised Volatility', f]:.0%})" for f in sorted_funds])
ws3.cell(row=label_row + 1, column=1, value=fund_labels).font = Font(
    name="Arial", size=9)

ws3.add_chart(scatter, "A" + str(label_row + 3))

set_widths(ws3, 4, default=20)


# ══════════════════════════════════════════════════════════════════════════
# SHEET 4: Calendar Year Returns Comparison
# ══════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Calendar Year Returns")
title_banner(ws4, "CALENDAR YEAR RETURNS — PEER GROUP", end_col=end_col, tab_color="E67E22")

years = sorted(df.index.year.unique())
row_h = 3
yr_headers = ["Year"] + FUND_NAMES
for c, h in enumerate(yr_headers, 1):
    ws4.cell(row=row_h, column=c, value=h)
style_header(ws4, row_h, len(yr_headers))

for yi, year in enumerate(years):
    r = row_h + 1 + yi
    ws4.cell(row=r, column=1, value=int(year)).number_format = "0"
    for fi, fund in enumerate(FUND_NAMES):
        val = yearly.loc[year, fund]
        cell = ws4.cell(row=r, column=fi + 2, value=val)
        cell.number_format = pct_fmt
        if val > 0:
            cell.font = Font(name="Arial", size=10, color=GREEN)
        elif val < 0:
            cell.font = Font(name="Arial", size=10, color=RED)
    style_row(ws4, r, len(yr_headers), alt=(yi % 2 == 0))

set_widths(ws4, len(yr_headers), default=16)

# Bar chart
n_yrs = len(years)
bar = BarChart()
bar.type = "col"
bar.title = "Calendar Year Returns"
bar.y_axis.numFmt = '0%'
bar.style = 10
bar.width = 32
bar.height = 16

for j in range(NUM_FUNDS):
    ref = Reference(ws4, min_col=j + 2, min_row=row_h, max_row=row_h + n_yrs)
    bar.add_data(ref, titles_from_data=True)
    bar.series[j].graphicalProperties.solidFill = CHART_COLOURS[j % len(CHART_COLOURS)]

cats = Reference(ws4, min_col=1, min_row=row_h + 1, max_row=row_h + n_yrs)
bar.set_categories(cats)
bar.legend.position = "b"
ws4.add_chart(bar, f"A{row_h + n_yrs + 3}")


# ══════════════════════════════════════════════════════════════════════════
# SHEET 5: Rolling Sharpe Comparison
# ══════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Rolling Sharpe")
title_banner(ws5, "ROLLING 12-MONTH SHARPE RATIO", end_col=end_col, tab_color="8E44AD")

row_h = 3
rs_headers = ["Date"] + FUND_NAMES
for c, h in enumerate(rs_headers, 1):
    ws5.cell(row=row_h, column=c, value=h)
style_header(ws5, row_h, len(rs_headers))

n_roll = len(rolling_sh)
for i, (dt, vals) in enumerate(rolling_sh.iterrows()):
    r = row_h + 1 + i
    ws5.cell(row=r, column=1, value=dt).number_format = "MMM-YY"
    for j, v in enumerate(vals):
        ws5.cell(row=r, column=j + 2, value=v).number_format = ratio_fmt

# Line chart
roll_chart = LineChart()
roll_chart.title = "Rolling 12-Month Sharpe Ratio"
roll_chart.y_axis.title = "Sharpe Ratio"
roll_chart.y_axis.numFmt = '0.0'
roll_chart.style = 10
roll_chart.width = 32
roll_chart.height = 16

for j in range(NUM_FUNDS):
    ref = Reference(ws5, min_col=j + 2, min_row=row_h, max_row=row_h + n_roll)
    roll_chart.add_data(ref, titles_from_data=True)
    roll_chart.series[j].graphicalProperties.line.width = 18000
    roll_chart.series[j].graphicalProperties.line.solidFill = CHART_COLOURS[j % len(CHART_COLOURS)]

cats = Reference(ws5, min_col=1, min_row=row_h + 1, max_row=row_h + n_roll)
roll_chart.set_categories(cats)
roll_chart.legend.position = "b"
ws5.add_chart(roll_chart, f"A{row_h + n_roll + 3}")
set_widths(ws5, len(rs_headers), default=14)


# ══════════════════════════════════════════════════════════════════════════
# SHEET 6: Cumulative Growth Comparison
# ══════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Cumulative Growth")
title_banner(ws6, "CUMULATIVE GROWTH OF $100 — PEER GROUP", end_col=end_col, tab_color=MED_BLUE)

prices = (1 + df).cumprod() * 100

row_h = 3
g_headers = ["Date"] + FUND_NAMES
for c, h in enumerate(g_headers, 1):
    ws6.cell(row=row_h, column=c, value=h)
style_header(ws6, row_h, len(g_headers))

for i, (dt, vals) in enumerate(prices.iterrows()):
    r = row_h + 1 + i
    ws6.cell(row=r, column=1, value=dt).number_format = "MMM-YY"
    for j, v in enumerate(vals):
        ws6.cell(row=r, column=j + 2, value=round(v, 2)).number_format = '$#,##0.00'

# Growth chart
g_chart = LineChart()
g_chart.title = "Cumulative Growth of $100"
g_chart.y_axis.title = "Value ($)"
g_chart.style = 10
g_chart.width = 32
g_chart.height = 16

for j in range(NUM_FUNDS):
    ref = Reference(ws6, min_col=j + 2, min_row=row_h, max_row=row_h + N)
    g_chart.add_data(ref, titles_from_data=True)
    g_chart.series[j].graphicalProperties.line.width = 22000
    g_chart.series[j].graphicalProperties.line.solidFill = CHART_COLOURS[j % len(CHART_COLOURS)]

cats = Reference(ws6, min_col=1, min_row=row_h + 1, max_row=row_h + N)
g_chart.set_categories(cats)
g_chart.legend.position = "b"
ws6.add_chart(g_chart, f"A{row_h + N + 3}")
set_widths(ws6, len(g_headers), default=14)


# ══════════════════════════════════════════════════════════════════════════
# SHEET 7: Drawdown Comparison
# ══════════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("Drawdown Comparison")
title_banner(ws7, "DRAWDOWN COMPARISON — PEER GROUP", end_col=end_col, tab_color=RED)

dd_df = pd.DataFrame()
for col in df.columns:
    cum = (1 + df[col]).cumprod()
    peak = cum.cummax()
    dd_df[col] = (cum - peak) / peak

row_h = 3
dd_headers = ["Date"] + FUND_NAMES
for c, h in enumerate(dd_headers, 1):
    ws7.cell(row=row_h, column=c, value=h)
style_header(ws7, row_h, len(dd_headers))

for i, (dt, vals) in enumerate(dd_df.iterrows()):
    r = row_h + 1 + i
    ws7.cell(row=r, column=1, value=dt).number_format = "MMM-YY"
    for j, v in enumerate(vals):
        ws7.cell(row=r, column=j + 2, value=v).number_format = pct_fmt

dd_chart = LineChart()
dd_chart.title = "Underwater Chart — Peer Group"
dd_chart.y_axis.title = "Drawdown"
dd_chart.y_axis.numFmt = '0%'
dd_chart.style = 10
dd_chart.width = 32
dd_chart.height = 16

for j in range(NUM_FUNDS):
    ref = Reference(ws7, min_col=j + 2, min_row=row_h, max_row=row_h + N)
    dd_chart.add_data(ref, titles_from_data=True)
    dd_chart.series[j].graphicalProperties.line.width = 18000
    dd_chart.series[j].graphicalProperties.line.solidFill = CHART_COLOURS[j % len(CHART_COLOURS)]

cats = Reference(ws7, min_col=1, min_row=row_h + 1, max_row=row_h + N)
dd_chart.set_categories(cats)
dd_chart.legend.position = "b"
ws7.add_chart(dd_chart, f"A{row_h + N + 3}")
set_widths(ws7, len(dd_headers), default=14)


# ══════════════════════════════════════════════════════════════════════════
# SHEET 8: Data
# ══════════════════════════════════════════════════════════════════════════
ws8 = wb.create_sheet("Data")
ws8.sheet_properties.tabColor = "95A5A6"
ws8["A1"] = "RAW MONTHLY RETURNS"
ws8["A1"].font = Font(name="Arial", bold=True, size=12, color=DARK_BLUE)

row_h = 2
d_headers = ["Date"] + FUND_NAMES
for c, h in enumerate(d_headers, 1):
    ws8.cell(row=row_h, column=c, value=h)
style_header(ws8, row_h, len(d_headers))

for i, (dt, vals) in enumerate(df.iterrows()):
    r = row_h + 1 + i
    ws8.cell(row=r, column=1, value=dt).number_format = "MMM-YY"
    for j, v in enumerate(vals):
        ws8.cell(row=r, column=j + 2, value=round(v, 6)).number_format = '0.0000%'

set_widths(ws8, len(d_headers), default=14)


# ── Save ──────────────────────────────────────────────────────────────────
OUTPUT = "Fund_Peer_Comparison.xlsx"
wb.save(OUTPUT)
print(f"✅ Peer comparison workbook saved: {OUTPUT}")
print(f"   Sheets: {wb.sheetnames}")
print(f"   Funds compared: {NUM_FUNDS}")
print(f"   Months: {N} ({period_start} to {period_end})")
print(f"   Overall ranking: {', '.join(f'{f} ({s:.1f})' for f, s in overall_rank_sorted.items())}")
